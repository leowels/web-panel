from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, desc, asc, nullslast, case
from sqlalchemy.orm import selectinload
from collections import defaultdict
from typing import Any, List, Optional, Union, Dict
from datetime import datetime, timedelta
from pydantic import BaseModel, ValidationError
import logging
import csv
import io
import re
import os
from pathlib import Path

from PIL import Image
import pytesseract

logger = logging.getLogger(__name__)

# Р СџР С•Р Т‘Р Т‘Р ВµРЎР‚Р В¶Р С”Р В° Р В·Р В°Р С—РЎС“РЎРѓР С”Р В° Р С”Р В°Р С” РЎРѓР С”РЎР‚Р С‘Р С—РЎвЂљР В° Р С‘ Р С”Р В°Р С” Р СР С•Р Т‘РЎС“Р В»РЎРЏ
try:
    from backend.models import Equipment, EquipmentHistory, UserActivity, User, UserRole, File, Violation
    from backend.database import get_db
    from backend.auth import get_current_user, require_permission
except ImportError:
    from ..models import Equipment, EquipmentHistory, UserActivity, User, UserRole, File, Violation
    from ..database import get_db
    from ..auth import get_current_user, require_permission

router = APIRouter(prefix="/api/equipment", tags=["equipment"])

EXPORT_STATUS_LABELS = {
    "active": "Активно",
    "inactive": "Неактивно",
    "archived": "Архив",
}


def _format_date_ru(value: Optional[datetime]) -> str:
    if not value:
        return ""
    return value.strftime("%d.%m.%Y")


def _format_bool_ru(value: Optional[bool]) -> str:
    return "Да" if bool(value) else "Нет"

DEFAULT_EQUIPMENT_TYPES = [
    "Кран",
    "Мостовой кран электрический (ЭМК)",
    "Электро мостовой кран (ЭМК)",
    "Кран-балка электрическая",
    "Кран-балка ручная",
    "Монорельс с электрической талью",
    "Кран консольный-поворотный",
    "Подъемник",
    "Лифт",
    "Эскалатор",
    "Другое",
]

class EquipmentCreate(BaseModel):
    equipment_type: str
    passport_number: str
    load_capacity: Optional[float] = None
    manufacturer: Optional[str] = None
    installation_date: Optional[datetime] = None
    pto_date: Optional[datetime] = None
    cto_date: Optional[datetime] = None
    installation_location: Optional[str] = None
    inventory_number: Optional[str] = None
    position: Optional[str] = None
    workshop: Optional[str] = None
    rostekhnadzor_registered: Optional[bool] = False
    expertise_date: Optional[datetime] = None
    operation_permit_until: Optional[datetime] = None
    operation_banned: Optional[bool] = False
    epb_positive_details: Optional[str] = None
    map_x: Optional[float] = None  # Р С™Р С•Р С•РЎР‚Р Т‘Р С‘Р Р…Р В°РЎвЂљР В° X Р Р…Р В° Р С”Р В°РЎР‚РЎвЂљР Вµ (0-100%)
    map_y: Optional[float] = None  # Р С™Р С•Р С•РЎР‚Р Т‘Р С‘Р Р…Р В°РЎвЂљР В° Y Р Р…Р В° Р С”Р В°РЎР‚РЎвЂљР Вµ (0-100%)
    status: Optional[str] = "active"

class EquipmentUpdate(BaseModel):
    equipment_type: Optional[str] = None
    passport_number: Optional[str] = None
    load_capacity: Optional[float] = None
    manufacturer: Optional[str] = None
    installation_date: Optional[datetime] = None
    pto_date: Optional[datetime] = None
    cto_date: Optional[datetime] = None
    installation_location: Optional[str] = None
    inventory_number: Optional[str] = None
    position: Optional[str] = None
    workshop: Optional[str] = None
    rostekhnadzor_registered: Optional[bool] = None
    expertise_date: Optional[datetime] = None
    operation_permit_until: Optional[datetime] = None
    operation_banned: Optional[bool] = None
    epb_positive_details: Optional[str] = None
    map_x: Optional[float] = None  # Р С™Р С•Р С•РЎР‚Р Т‘Р С‘Р Р…Р В°РЎвЂљР В° X Р Р…Р В° Р С”Р В°РЎР‚РЎвЂљР Вµ (0-100%)
    map_y: Optional[float] = None  # Р С™Р С•Р С•РЎР‚Р Т‘Р С‘Р Р…Р В°РЎвЂљР В° Y Р Р…Р В° Р С”Р В°РЎР‚РЎвЂљР Вµ (0-100%)
    status: Optional[str] = None

class EquipmentResponse(BaseModel):
    id: int
    equipment_type: str
    passport_number: str
    inventory_number: Optional[str]
    position: Optional[str]
    workshop: Optional[str]
    rostekhnadzor_registered: Optional[bool]
    expertise_date: Optional[datetime]
    operation_permit_until: Optional[datetime]
    operation_banned: Optional[bool]
    epb_positive_details: Optional[str]
    map_x: Optional[float]  # Р С™Р С•Р С•РЎР‚Р Т‘Р С‘Р Р…Р В°РЎвЂљР В° X Р Р…Р В° Р С”Р В°РЎР‚РЎвЂљР Вµ (0-100%)
    map_y: Optional[float]  # Р С™Р С•Р С•РЎР‚Р Т‘Р С‘Р Р…Р В°РЎвЂљР В° Y Р Р…Р В° Р С”Р В°РЎР‚РЎвЂљР Вµ (0-100%)
    load_capacity: Optional[float]
    manufacturer: Optional[str]
    installation_date: Optional[datetime]
    pto_date: Optional[datetime]
    cto_date: Optional[datetime]
    installation_location: Optional[str]
    status: str
    created_at: datetime
    updated_at: datetime
    violations_open: Optional[int] = 0
    violations_total: Optional[int] = 0

    class Config:
        from_attributes = True
class EquipmentListResponse(BaseModel):
    items: List[EquipmentResponse]
    total: int

class EquipmentBulkItem(BaseModel):
    equipment_type: str
    passport_number: str
    load_capacity: Optional[float] = None
    manufacturer: Optional[str] = None
    installation_date: Optional[datetime] = None
    pto_date: Optional[datetime] = None
    cto_date: Optional[datetime] = None
    installation_location: Optional[str] = None
    inventory_number: Optional[str] = None
    position: Optional[str] = None
    workshop: Optional[str] = None
    rostekhnadzor_registered: Optional[bool] = False
    expertise_date: Optional[datetime] = None
    operation_permit_until: Optional[datetime] = None
    operation_banned: Optional[bool] = False
    epb_positive_details: Optional[str] = None
    status: Optional[str] = "active"

class EquipmentBulkRequest(BaseModel):
    items: List[EquipmentBulkItem]
    skip_duplicates: bool = True

class EquipmentBulkResponse(BaseModel):
    created: int
    skipped: int
    created_ids: List[int]
    errors: List[dict]

    class Config:
        arbitrary_types_allowed = True
        from_attributes = True

class EquipmentBulkUpdateRequest(BaseModel):
    equipment_ids: List[int]
    update_data: EquipmentUpdate

class EquipmentBulkUpdateResponse(BaseModel):
    updated: int
    errors: List[dict]

class EquipmentBulkDatesRequest(BaseModel):
    equipment_ids: List[int]
    pto_date: Optional[datetime] = None
    cto_date: Optional[datetime] = None

class EquipmentBulkDatesResponse(BaseModel):
    updated: int
    errors: List[dict]

class EquipmentOCRUpsertRequest(BaseModel):
    name: Optional[str] = None
    capacity: Optional[float] = None
    inventory_number: Optional[str] = None
    manufacturer: Optional[str] = None
    installation_date: Optional[datetime] = None
    pto_date: Optional[datetime] = None
    cto_date: Optional[datetime] = None
    equipment_type: Optional[str] = None
    passport_number: Optional[str] = None
    position: Optional[str] = None
    workshop: Optional[str] = None

class EquipmentOCRUpsertResponse(BaseModel):
    id: int
    created: bool  # True Р ВµРЎРѓР В»Р С‘ РЎРѓР С•Р В·Р Т‘Р В°Р Р… Р Р…Р С•Р Р†РЎвЂ№Р в„–, False Р ВµРЎРѓР В»Р С‘ Р С•Р В±Р Р…Р С•Р Р†Р В»Р ВµР Р… РЎРѓРЎС“РЎвЂ°Р ВµРЎРѓРЎвЂљР Р†РЎС“РЎР‹РЎвЂ°Р С‘Р в„–


class EquipmentOCRImportRequest(BaseModel):
    """
    Р вЂ”Р В°Р С—РЎР‚Р С•РЎРѓ Р Р…Р В° Р С‘Р СР С—Р С•РЎР‚РЎвЂљ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ РЎвЂЎР ВµРЎР‚Р ВµР В· OCR/РЎвЂљР В°Р В±Р В»Р С‘РЎвЂЎР Р…РЎвЂ№Р в„– РЎвЂљР ВµР С”РЎРѓРЎвЂљ.
    Р вЂ™Р В°РЎР‚Р С‘Р В°Р Р…РЎвЂљРЎвЂ№:
    - ocr_text: РЎС“Р В¶Р Вµ РЎР‚Р В°РЎРѓР С—Р С•Р В·Р Р…Р В°Р Р…Р Р…РЎвЂ№Р в„– РЎвЂљР ВµР С”РЎРѓРЎвЂљ РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ РЎвЂ№ (CSV-Р С—Р С•Р Т‘Р С•Р В±Р Р…РЎвЂ№Р в„–)
    - file_id: ID РЎвЂћР В°Р в„–Р В»Р В° Р Р† РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ Р Вµ files (РЎвЂћР С•РЎвЂљР С• Р С‘Р В»Р С‘ CSV/РЎвЂљР ВµР С”РЎРѓРЎвЂљ)
    """
    ocr_text: Optional[str] = None
    file_id: Optional[int] = None


class EquipmentRiskResponse(BaseModel):
    equipment_id: int
    risk_score: float
    risk_level: str
    active_violations: int
    overdue: int
    repeat_violations: int


class EquipmentRiskTopItem(BaseModel):
    equipment_id: int
    passport_number: str
    equipment_type: str
    workshop: Optional[str] = None
    risk_score: float
    risk_level: str
    active_violations: int
    overdue: int
    repeat_violations: int


class EquipmentRiskTopResponse(BaseModel):
    items: List[EquipmentRiskTopItem]


def _equipment_to_response(
    equipment: Equipment,
    violations_open: int = 0,
    violations_total: int = 0,
) -> EquipmentResponse:
    return EquipmentResponse(
        id=equipment.id,
        equipment_type=equipment.equipment_type,
        passport_number=equipment.passport_number,
        inventory_number=equipment.inventory_number,
        position=equipment.position,
        workshop=equipment.workshop,
        rostekhnadzor_registered=equipment.rostekhnadzor_registered,
        expertise_date=equipment.expertise_date,
        operation_permit_until=equipment.operation_permit_until,
        operation_banned=equipment.operation_banned,
        epb_positive_details=equipment.epb_positive_details,
        map_x=equipment.map_x,
        map_y=equipment.map_y,
        load_capacity=equipment.load_capacity,
        manufacturer=equipment.manufacturer,
        installation_date=equipment.installation_date,
        pto_date=equipment.pto_date,
        cto_date=equipment.cto_date,
        installation_location=equipment.installation_location,
        status=equipment.status,
        created_at=equipment.created_at,
        updated_at=equipment.updated_at,
        violations_open=violations_open,
        violations_total=violations_total,
    )


def _build_equipment_filters(
    search: Optional[str],
    equipment_type: Optional[str],
    status: Optional[str],
    workshop: Optional[str],
    maintenance: Optional[str],
    maintenance_scope: Optional[str],
):
    filters = []

    if search:
        filters.append(
            or_(
                Equipment.passport_number.ilike(f"%{search}%"),
                Equipment.inventory_number.ilike(f"%{search}%"),
                Equipment.position.ilike(f"%{search}%"),
                Equipment.equipment_type.ilike(f"%{search}%"),
                Equipment.workshop.ilike(f"%{search}%"),
                Equipment.installation_location.ilike(f"%{search}%"),
            )
        )
    if equipment_type:
        filters.append(Equipment.equipment_type == equipment_type)
    if status:
        filters.append(Equipment.status == status)
    if workshop:
        filters.append(Equipment.workshop == workshop)

    if maintenance:
        now = datetime.utcnow()
        if maintenance == "overdue":
            def _cond(col):
                return and_(col.isnot(None), col < now)
        elif maintenance == "due_30":
            end = now + timedelta(days=30)
            def _cond(col):
                return and_(col.isnot(None), col >= now, col <= end)
        else:
            end = now + timedelta(days=60)
            def _cond(col):
                return and_(col.isnot(None), col >= now, col <= end)

        if maintenance_scope == "pto":
            filters.append(_cond(Equipment.pto_date))
        elif maintenance_scope == "cto":
            filters.append(_cond(Equipment.cto_date))
        else:
            filters.append(or_(_cond(Equipment.pto_date), _cond(Equipment.cto_date)))

    return filters


def _apply_equipment_sort(query, sort_by: Optional[str], sort_dir: Optional[str]):
    sort_map = {
        "updated_at": Equipment.updated_at,
        "passport_number": Equipment.passport_number,
        "equipment_type": Equipment.equipment_type,
        "status": Equipment.status,
        "pto_date": Equipment.pto_date,
        "cto_date": Equipment.cto_date,
        "installation_date": Equipment.installation_date,
    }
    sort_col = sort_map.get(sort_by or "updated_at", Equipment.updated_at)
    sort_func = desc if (sort_dir or "desc") == "desc" else asc
    if sort_by in ("pto_date", "cto_date", "installation_date"):
        return query.order_by(nullslast(sort_func(sort_col)))
    return query.order_by(sort_func(sort_col))


SEVERITY_WEIGHT = {
    "low": 1.0,
    "medium": 2.0,
    "high": 3.0,
    "critical": 5.0,
}


def _risk_level_from_score(score: float) -> str:
    if score >= 25:
        return "critical"
    if score >= 16:
        return "high"
    if score >= 6:
        return "medium"
    return "low"


def _build_repeat_key(violation: Any) -> str:
    violation_type = (getattr(violation, "violation_type", None) or "").strip().lower()
    if violation_type:
        return f"type:{violation_type}"
    description = (getattr(violation, "description", None) or "").strip().lower()
    if description:
        return f"desc:{description[:120]}"
    return f"id:{getattr(violation, 'id', 'unknown')}"


def _calculate_equipment_risk(violations: List[Any]) -> Dict[str, Any]:
    now = datetime.utcnow()
    active_violations = len(violations)
    overdue = 0
    weighted_sum = 0.0
    repeat_counter: Dict[str, int] = defaultdict(int)

    for violation in violations:
        severity = (getattr(violation, "severity", None) or "medium").lower()
        weighted_sum += SEVERITY_WEIGHT.get(severity, SEVERITY_WEIGHT["medium"])

        deadline = getattr(violation, "deadline", None)
        if deadline and deadline < now:
            overdue += 1

        repeat_key = _build_repeat_key(violation)
        repeat_counter[repeat_key] += 1

    repeat_violations = sum(count for count in repeat_counter.values() if count > 1)
    risk_score = weighted_sum + (overdue * 2.0) + (repeat_violations * 1.5)
    risk_score = round(risk_score, 1)

    return {
        "risk_score": risk_score,
        "risk_level": _risk_level_from_score(risk_score),
        "active_violations": active_violations,
        "overdue": overdue,
        "repeat_violations": repeat_violations,
    }


async def _bulk_create_equipment_items(
    items: List[EquipmentBulkItem],
    skip_duplicates: bool,
    current_user: User,
    db: AsyncSession
) -> EquipmentBulkResponse:
    created_ids: List[int] = []
    skipped = 0
    errors: List[dict] = []

    for index, item in enumerate(items):
        try:
            # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С Р С•Р В±РЎРЏР В·Р В°РЎвЂљР ВµР В»РЎРЉР Р…РЎвЂ№Р Вµ Р С—Р С•Р В»РЎРЏ
            if not item.passport_number or not item.equipment_type:
                raise HTTPException(status_code=400, detail="Passport number and equipment type are required")

            # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С Р Т‘РЎС“Р В±Р В»Р С‘Р С”Р В°РЎвЂљРЎвЂ№ Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљР В°
            existing_passport = await db.execute(
                select(Equipment.id).where(Equipment.passport_number == item.passport_number)
            )
            if existing_passport.scalar_one_or_none():
                if skip_duplicates:
                    errors.append(
                        {
                            "index": index,
                            "passport_number": item.passport_number,
                            "detail": "Passport number already exists",
                            "type": "duplicate",
                        }
                    )
                    skipped += 1
                    continue
                raise HTTPException(status_code=400, detail="Passport number already exists")

            # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С Р Т‘РЎС“Р В±Р В»Р С‘Р С”Р В°РЎвЂљРЎвЂ№ Р С‘Р Р…Р Р†Р ВµР Р…РЎвЂљР В°РЎР‚Р Р…Р С•Р С–Р С• Р Р…Р С•Р СР ВµРЎР‚Р В°
            if item.inventory_number:
                existing_inventory = await db.execute(
                    select(Equipment.id).where(Equipment.inventory_number == item.inventory_number)
                )
                if existing_inventory.scalar_one_or_none():
                    if skip_duplicates:
                        errors.append(
                            {
                                "index": index,
                                "passport_number": item.passport_number,
                                "detail": "Inventory number already exists",
                                "type": "duplicate",
                            }
                        )
                        skipped += 1
                        continue
                    raise HTTPException(status_code=400, detail="Inventory number already exists")

            new_equipment = Equipment(
                **item.dict(),
                created_by=current_user.id,
                status=item.status or "active"
            )
            db.add(new_equipment)
            await db.flush()
            created_ids.append(new_equipment.id)

            activity = UserActivity(
                user_id=current_user.id,
                action_type="create",
                entity_type="equipment",
                entity_id=new_equipment.id,
                description=f"Bulk created equipment {new_equipment.passport_number}"
            )
            db.add(activity)
        except HTTPException as http_exc:
            errors.append(
                {
                    "index": index,
                    "passport_number": item.passport_number,
                    "detail": http_exc.detail,
                }
            )
            if not skip_duplicates:
                await db.rollback()
                raise
        except Exception as exc:
            errors.append(
                {
                    "index": index,
                    "passport_number": item.passport_number,
                    "detail": str(exc),
                }
            )

    await db.commit()

    return EquipmentBulkResponse(
        created=len(created_ids),
        skipped=skipped,
        created_ids=created_ids,
        errors=errors,
    )


def _normalize_csv_date(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    # Р СџР С•Р Т‘Р Т‘Р ВµРЎР‚Р В¶Р С‘Р Р†Р В°Р ВµР С РЎвЂћР С•РЎР‚Р СР В°РЎвЂљРЎвЂ№ YYYY-MM-DD Р С‘ DD.MM.YYYY
    if re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return f"{value}T00:00:00"
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", value):
        dt = datetime.strptime(value, "%d.%m.%Y")
        return dt.strftime("%Y-%m-%dT00:00:00")
    # Р С›РЎРѓРЎвЂљР В°Р Р†Р В»РЎРЏР ВµР С Р С”Р В°Р С” Р ВµРЎРѓРЎвЂљРЎРЉ - Pydantic Р С—Р С•Р С—РЎР‚Р С•Р В±РЎС“Р ВµРЎвЂљ РЎР‚Р В°РЎРѓР С—Р В°РЎР‚РЎРѓР С‘РЎвЂљРЎРЉ
    return value


def _normalize_float(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    value = value.strip().replace(",", ".")
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _normalize_bool(value: Optional[str]) -> Optional[bool]:
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if not normalized:
        return None
    if normalized in {"1", "true", "yes", "да", "y"}:
        return True
    if normalized in {"0", "false", "no", "нет", "n"}:
        return False
    return None


def _ocr_image_to_text(image_path: str) -> str:
    """
    Р СџРЎР‚Р ВµР С•Р В±РЎР‚Р В°Р В·Р С•Р Р†Р В°Р Р…Р С‘Р Вµ Р С‘Р В·Р С•Р В±РЎР‚Р В°Р В¶Р ВµР Р…Р С‘РЎРЏ (РЎвЂћР С•РЎвЂљР С• РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ РЎвЂ№) Р Р† РЎвЂљР ВµР С”РЎРѓРЎвЂљ РЎРѓ Р С—Р С•Р СР С•РЎвЂ°РЎРЉРЎР‹ Tesseract OCR.
    Р ВРЎРѓР С—Р С•Р В»РЎРЉР В·РЎС“Р ВµРЎвЂљРЎРѓРЎРЏ Р С”Р В°Р С” Р В±Р ВµРЎРѓР С—Р В»Р В°РЎвЂљР Р…РЎвЂ№Р в„– Р В»Р С•Р С”Р В°Р В»РЎРЉР Р…РЎвЂ№Р в„– OCR.
    """
    try:
        # Р С›РЎвЂљР С”РЎР‚РЎвЂ№Р Р†Р В°Р ВµР С Р С‘Р В·Р С•Р В±РЎР‚Р В°Р В¶Р ВµР Р…Р С‘Р Вµ
        img = Image.open(image_path)
        logger.info(f"Opened image: {img.format}, size: {img.size}, mode: {img.mode}")
        
        # Р С™Р С•Р Р…Р Р†Р ВµРЎР‚РЎвЂљР С‘РЎР‚РЎС“Р ВµР С Р Р† RGB, Р ВµРЎРѓР В»Р С‘ Р Р…РЎС“Р В¶Р Р…Р С• (Р Т‘Р В»РЎРЏ РЎвЂћР С•РЎР‚Р СР В°РЎвЂљР С•Р Р† РЎРѓ Р С—РЎР‚Р С•Р В·РЎР‚Р В°РЎвЂЎР Р…Р С•РЎРѓРЎвЂљРЎРЉРЎР‹ Р С‘Р В»Р С‘ Р Т‘РЎР‚РЎС“Р С–Р С‘РЎвЂ¦ РЎР‚Р ВµР В¶Р С‘Р СР С•Р Р†)
        if img.mode != 'RGB':
            logger.info(f"Converting image from {img.mode} to RGB")
            # Р РЋР С•Р В·Р Т‘Р В°Р ВµР С Р В±Р ВµР В»РЎвЂ№Р в„– РЎвЂћР С•Р Р… Р Т‘Р В»РЎРЏ Р С‘Р В·Р С•Р В±РЎР‚Р В°Р В¶Р ВµР Р…Р С‘Р в„– РЎРѓ Р С—РЎР‚Р С•Р В·РЎР‚Р В°РЎвЂЎР Р…Р С•РЎРѓРЎвЂљРЎРЉРЎР‹
            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'RGBA':
                rgb_img.paste(img, mask=img.split()[3])  # Р ВРЎРѓР С—Р С•Р В»РЎРЉР В·РЎС“Р ВµР С Р В°Р В»РЎРЉРЎвЂћР В°-Р С”Р В°Р Р…Р В°Р В» Р С”Р В°Р С” Р СР В°РЎРѓР С”РЎС“
            else:
                rgb_img.paste(img)
            img = rgb_img
        
        # Р Р€Р В±Р ВµР В¶Р Т‘Р В°Р ВµР СРЎРѓРЎРЏ, РЎвЂЎРЎвЂљР С• Р С‘Р В·Р С•Р В±РЎР‚Р В°Р В¶Р ВµР Р…Р С‘Р Вµ Р Р† Р С—РЎР‚Р В°Р Р†Р С‘Р В»РЎРЉР Р…Р С•Р С РЎвЂћР С•РЎР‚Р СР В°РЎвЂљР Вµ Р Т‘Р В»РЎРЏ pytesseract
        if img.mode != 'RGB':
            img = img.convert('RGB')
            
    except Exception as e:
        logger.error(f"Failed to open/process image for OCR: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to open image for OCR: {str(e)}")

    try:
        # Р В РЎС“РЎРѓРЎРѓР С”Р С‘Р в„– + Р В°Р Р…Р С–Р В»Р С‘Р в„–РЎРѓР С”Р С‘Р в„–, РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ РЎвЂ№ Р Р…Р В° Р СџР РЋ Р С•Р В±РЎвЂ№РЎвЂЎР Р…Р С• Р Р…Р В° РЎР‚РЎС“РЎРѓРЎРѓР С”Р С•Р С
        # Р ВРЎРѓР С—Р С•Р В»РЎРЉР В·РЎС“Р ВµР С Р Т‘Р С•Р С—Р С•Р В»Р Р…Р С‘РЎвЂљР ВµР В»РЎРЉР Р…РЎС“РЎР‹ Р С•Р В±РЎР‚Р В°Р В±Р С•РЎвЂљР С”РЎС“ Р Т‘Р В»РЎРЏ Р В»РЎС“РЎвЂЎРЎв‚¬Р ВµР С–Р С• РЎР‚Р В°РЎРѓР С—Р С•Р В·Р Р…Р В°Р Р†Р В°Р Р…Р С‘РЎРЏ РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ 
        text = pytesseract.image_to_string(img, lang="rus+eng", config='--psm 6')
        logger.info(f"OCR completed, extracted {len(text)} characters")
        return text
    except Exception as e:
        logger.error(f"OCR error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"OCR error: {str(e)}")


@router.get("", response_model=Union[List[EquipmentResponse], EquipmentListResponse])
async def get_equipment_list(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    page: Optional[int] = Query(None, ge=1),
    page_size: Optional[int] = Query(None, ge=1, le=1000),
    search: Optional[str] = None,
    equipment_type: Optional[str] = None,
    status: Optional[str] = None,
    workshop: Optional[str] = None,
    maintenance: Optional[str] = Query(None, regex="^(overdue|due_30|due_60)$"),
    maintenance_scope: Optional[str] = Query("any", regex="^(any|pto|cto)$"),
    sort_by: Optional[str] = Query("updated_at", regex="^(updated_at|passport_number|equipment_type|status|pto_date|cto_date|installation_date)$"),
    sort_dir: Optional[str] = Query("desc", regex="^(asc|desc)$"),
    with_total: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СџР С•Р В»РЎС“РЎвЂЎР С‘РЎвЂљРЎРЉ РЎРѓР С—Р С‘РЎРѓР С•Р С” Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ"""
    await require_permission(current_user, "equipment:read", db)
    
    if page is not None:
        limit = page_size or limit
        skip = (page - 1) * limit
    elif page_size is not None:
        limit = page_size

    filters = _build_equipment_filters(
        search=search,
        equipment_type=equipment_type,
        status=status,
        workshop=workshop,
        maintenance=maintenance,
        maintenance_scope=maintenance_scope,
    )

    query = select(Equipment)
    if filters:
        query = query.where(and_(*filters))

    query = _apply_equipment_sort(query, sort_by, sort_dir)

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    equipment_list = result.scalars().all()

    equipment_ids = [eq.id for eq in equipment_list]
    violations_map: dict[int, dict[str, int]] = {}
    if equipment_ids:
        counts_result = await db.execute(
            select(
                Violation.equipment_id,
                func.count().label("total"),
                func.sum(case((Violation.status != "resolved", 1), else_=0)).label("open"),
            )
            .where(Violation.equipment_id.in_(equipment_ids))
            .group_by(Violation.equipment_id)
        )
        for row in counts_result.all():
            if not row:
                continue
            violations_map[row[0]] = {
                "total": int(row[1] or 0),
                "open": int(row[2] or 0),
            }

    items = [
        _equipment_to_response(
            eq,
            violations_open=violations_map.get(eq.id, {}).get("open", 0),
            violations_total=violations_map.get(eq.id, {}).get("total", 0),
        )
        for eq in equipment_list
    ]
    if with_total:
        count_query = select(func.count()).select_from(Equipment)
        if filters:
            count_query = count_query.where(and_(*filters))
        total = (await db.execute(count_query)).scalar() or 0
        return EquipmentListResponse(items=items, total=total)
    return items

@router.get("/types", response_model=List[str])
async def get_equipment_types(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СџР С•Р В»РЎС“РЎвЂЎР С‘РЎвЂљРЎРЉ РЎРѓР С—Р С‘РЎРѓР С•Р С” РЎвЂљР С‘Р С—Р С•Р Р† Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ"""
    await require_permission(current_user, "equipment:read", db)
    result = await db.execute(
        select(Equipment.equipment_type)
        .where(Equipment.equipment_type.isnot(None))
        .distinct()
        .order_by(Equipment.equipment_type)
    )
    db_types = [row[0] for row in result.all() if row and row[0]]

    # Always include baseline catalogue types, then append custom DB values.
    merged: List[str] = []
    seen = set()
    for item in DEFAULT_EQUIPMENT_TYPES + sorted(db_types):
        if not item:
            continue
        key = item.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(key)
    return merged


@router.get("/export")
async def export_equipment_csv(
    search: Optional[str] = None,
    equipment_type: Optional[str] = None,
    status: Optional[str] = None,
    workshop: Optional[str] = None,
    maintenance: Optional[str] = Query(None, regex="^(overdue|due_30|due_60)$"),
    maintenance_scope: Optional[str] = Query("any", regex="^(any|pto|cto)$"),
    export_format: str = Query("xlsx", alias="format", regex="^(csv|xlsx)$"),
    sort_by: Optional[str] = Query("updated_at", regex="^(updated_at|passport_number|equipment_type|status|pto_date|cto_date|installation_date)$"),
    sort_dir: Optional[str] = Query("desc", regex="^(asc|desc)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Экспорт оборудования в CSV или XLSX c русскими полями и визуальным форматированием."""
    await require_permission(current_user, "equipment:read", db)

    filters = _build_equipment_filters(
        search=search,
        equipment_type=equipment_type,
        status=status,
        workshop=workshop,
        maintenance=maintenance,
        maintenance_scope=maintenance_scope,
    )

    query = select(Equipment)
    if filters:
        query = query.where(and_(*filters))
    query = _apply_equipment_sort(query, sort_by, sort_dir)

    result = await db.execute(query)
    equipment_list = result.scalars().all()

    violation_stats: Dict[int, Dict[str, int]] = {}
    equipment_ids = [eq.id for eq in equipment_list]
    if equipment_ids:
        stats_result = await db.execute(
            select(
                Violation.equipment_id,
                func.count(Violation.id).label("total_count"),
                func.sum(
                    case((Violation.status != "resolved", 1), else_=0)
                ).label("open_count"),
            )
            .where(Violation.equipment_id.in_(equipment_ids))
            .group_by(Violation.equipment_id)
        )
        for row in stats_result:
            violation_stats[int(row.equipment_id)] = {
                "open": int(row.open_count or 0),
                "total": int(row.total_count or 0),
            }

    columns = [
        ("id", "ID"),
        ("equipment_type", "Тип оборудования"),
        ("passport_number", "Паспорт"),
        ("inventory_number", "Инвентарный №"),
        ("workshop", "Цех"),
        ("position", "Позиция"),
        ("status", "Статус"),
        ("load_capacity", "Грузоподъемность, т"),
        ("manufacturer", "Производитель"),
        ("installation_location", "Место установки"),
        ("installation_date", "Дата ввода"),
        ("pto_date", "Дата ПТО"),
        ("cto_date", "Дата ЧТО"),
        ("rostekhnadzor_registered", "Регистрация в Ростехнадзоре"),
        ("expertise_date", "Дата экспертизы"),
        ("operation_permit_until", "Разрешено до"),
        ("operation_banned", "Запрет эксплуатации"),
        ("epb_positive_details", "Реквизиты положительной ЭПБ"),
        ("violations_open", "Нарушений в работе"),
        ("violations_total", "Нарушений всего"),
        ("updated_at", "Обновлено"),
    ]

    rows = []
    for eq in equipment_list:
        stats = violation_stats.get(eq.id, {"open": 0, "total": 0})
        rows.append(
            {
                "id": eq.id,
                "equipment_type": eq.equipment_type or "",
                "passport_number": eq.passport_number or "",
                "inventory_number": eq.inventory_number or "",
                "workshop": eq.workshop or "",
                "position": eq.position or "",
                "status": EXPORT_STATUS_LABELS.get(eq.status or "", eq.status or ""),
                "load_capacity": eq.load_capacity if eq.load_capacity is not None else "",
                "manufacturer": eq.manufacturer or "",
                "installation_location": eq.installation_location or "",
                "installation_date": _format_date_ru(eq.installation_date),
                "pto_date": _format_date_ru(eq.pto_date),
                "cto_date": _format_date_ru(eq.cto_date),
                "rostekhnadzor_registered": _format_bool_ru(eq.rostekhnadzor_registered),
                "expertise_date": _format_date_ru(eq.expertise_date),
                "operation_permit_until": _format_date_ru(eq.operation_permit_until),
                "operation_banned": _format_bool_ru(eq.operation_banned),
                "epb_positive_details": eq.epb_positive_details or "",
                "violations_open": stats["open"],
                "violations_total": stats["total"],
                "updated_at": _format_date_ru(eq.updated_at),
            }
        )

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if export_format == "csv":
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";")
        writer.writerow([title for _, title in columns])
        for row in rows:
            writer.writerow([row[key] for key, _ in columns])
        filename = f"equipment_{timestamp}.csv"
        return Response(
            content=output.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Для XLSX-экспорта требуется пакет openpyxl",
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Оборудование"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    status_fills = {
        "Активно": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "Неактивно": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Архив": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
    }

    for col_idx, (_, title) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    widths = [8, 30, 16, 18, 14, 14, 14, 18, 20, 24, 14, 14, 14, 20, 16, 16, 16, 40, 18, 16, 14]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    status_col_idx = [i for i, (key, _) in enumerate(columns, start=1) if key == "status"][0]
    register_col_idx = [i for i, (key, _) in enumerate(columns, start=1) if key == "rostekhnadzor_registered"][0]
    ban_col_idx = [i for i, (key, _) in enumerate(columns, start=1) if key == "operation_banned"][0]

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = border
            cell.alignment = body_alignment
            if row_idx % 2 == 0:
                cell.fill = even_fill

        status_value = row_data["status"]
        if status_value in status_fills:
            sheet.cell(row=row_idx, column=status_col_idx).fill = status_fills[status_value]

        register_cell = sheet.cell(row=row_idx, column=register_col_idx)
        register_cell.fill = PatternFill(
            start_color="DCFCE7" if row_data["rostekhnadzor_registered"] == "Да" else "FEE2E2",
            end_color="DCFCE7" if row_data["rostekhnadzor_registered"] == "Да" else "FEE2E2",
            fill_type="solid",
        )

        ban_cell = sheet.cell(row=row_idx, column=ban_col_idx)
        ban_cell.fill = PatternFill(
            start_color="FEE2E2" if row_data["operation_banned"] == "Да" else "DCFCE7",
            end_color="FEE2E2" if row_data["operation_banned"] == "Да" else "DCFCE7",
            fill_type="solid",
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:U{max(2, len(rows) + 1)}"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"equipment_{timestamp}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/risk/top", response_model=EquipmentRiskTopResponse)
async def get_top_risk_equipment(
    limit: int = Query(5, ge=1, le=50),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Топ оборудования с максимальным risk_score."""
    await require_permission(current_user, "equipment:read", db)

    equipment_result = await db.execute(
        select(Equipment).where(Equipment.status != "archived")
    )
    equipment_list = equipment_result.scalars().all()
    if not equipment_list:
        return EquipmentRiskTopResponse(items=[])

    equipment_ids = [eq.id for eq in equipment_list]
    violations_result = await db.execute(
        select(Violation).where(
            Violation.equipment_id.in_(equipment_ids),
            Violation.status != "resolved",
        )
    )
    open_violations = violations_result.scalars().all()

    violations_by_equipment: Dict[int, List[Violation]] = defaultdict(list)
    for violation in open_violations:
        violations_by_equipment[violation.equipment_id].append(violation)

    items: List[EquipmentRiskTopItem] = []
    for equipment in equipment_list:
        metrics = _calculate_equipment_risk(violations_by_equipment.get(equipment.id, []))
        items.append(
            EquipmentRiskTopItem(
                equipment_id=equipment.id,
                passport_number=equipment.passport_number,
                equipment_type=equipment.equipment_type,
                workshop=equipment.workshop,
                risk_score=metrics["risk_score"],
                risk_level=metrics["risk_level"],
                active_violations=metrics["active_violations"],
                overdue=metrics["overdue"],
                repeat_violations=metrics["repeat_violations"],
            )
        )

    items.sort(key=lambda item: item.risk_score, reverse=True)
    return EquipmentRiskTopResponse(items=items[:limit])


@router.get("/{equipment_id}/risk", response_model=EquipmentRiskResponse)
async def get_equipment_risk(
    equipment_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Расчет risk_score для конкретного оборудования."""
    await require_permission(current_user, "equipment:read", db)

    equipment_result = await db.execute(
        select(Equipment.id).where(Equipment.id == equipment_id)
    )
    if equipment_result.scalar_one_or_none() is None:
        raise HTTPException(status_code=404, detail="Equipment not found")

    violations_result = await db.execute(
        select(Violation).where(
            Violation.equipment_id == equipment_id,
            Violation.status != "resolved",
        )
    )
    open_violations = violations_result.scalars().all()
    metrics = _calculate_equipment_risk(open_violations)

    return EquipmentRiskResponse(
        equipment_id=equipment_id,
        risk_score=metrics["risk_score"],
        risk_level=metrics["risk_level"],
        active_violations=metrics["active_violations"],
        overdue=metrics["overdue"],
        repeat_violations=metrics["repeat_violations"],
    )

@router.get("/{equipment_id}", response_model=EquipmentResponse)
async def get_equipment(
    equipment_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СџР С•Р В»РЎС“РЎвЂЎР С‘РЎвЂљРЎРЉ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ Р С—Р С• ID"""
    await require_permission(current_user, "equipment:read", db)
    
    result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equipment = result.scalar_one_or_none()
    
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    
    violations_total = (await db.execute(
        select(func.count()).select_from(Violation).where(Violation.equipment_id == equipment_id)
    )).scalar() or 0
    violations_open = (await db.execute(
        select(func.count()).select_from(Violation).where(
            Violation.equipment_id == equipment_id,
            Violation.status != "resolved",
        )
    )).scalar() or 0

    return _equipment_to_response(
        equipment,
        violations_open=violations_open,
        violations_total=violations_total,
    )

@router.post("", response_model=EquipmentResponse, status_code=status.HTTP_201_CREATED)
async def create_equipment(
    equipment_data: EquipmentCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р РЋР С•Р В·Р Т‘Р В°РЎвЂљРЎРЉ Р Р…Р С•Р Р†Р С•Р Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ"""
    try:
        logger.debug(f"Creating equipment for user {current_user.id}")
        
        # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚Р С”Р В° Р С—РЎР‚Р В°Р Р† - Р С‘РЎРѓР С—Р С•Р В»РЎРЉР В·РЎС“Р ВµР С РЎС“Р В¶Р Вµ Р В·Р В°Р С–РЎР‚РЎС“Р В¶Р ВµР Р…Р Р…РЎвЂ№Р Вµ РЎР‚Р С•Р В»Р С‘ Р С‘Р В· get_current_user
        # get_current_user РЎС“Р В¶Р Вµ Р В·Р В°Р С–РЎР‚РЎС“Р В¶Р В°Р ВµРЎвЂљ РЎР‚Р С•Р В»Р С‘ РЎвЂЎР ВµРЎР‚Р ВµР В· selectinload
        user_roles = []
        try:
            # Р СџРЎР‚Р С•Р В±РЎС“Р ВµР С Р С—Р С•Р В»РЎС“РЎвЂЎР С‘РЎвЂљРЎРЉ РЎР‚Р С•Р В»Р С‘ Р С‘Р В· РЎС“Р В¶Р Вµ Р В·Р В°Р С–РЎР‚РЎС“Р В¶Р ВµР Р…Р Р…Р С•Р С–Р С• Р С•Р В±РЎР‰Р ВµР С”РЎвЂљР В°
            if hasattr(current_user, 'roles') and current_user.roles:
                user_roles = [ur.role.name for ur in current_user.roles]
        except Exception as e:
            logger.warning(f"Error getting roles from current_user: {e}")
            # Р вЂўРЎРѓР В»Р С‘ Р Р…Р Вµ Р С—Р С•Р В»РЎС“РЎвЂЎР С‘Р В»Р С•РЎРѓРЎРЉ, Р В·Р В°Р С–РЎР‚РЎС“Р В¶Р В°Р ВµР С Р В·Р В°Р Р…Р С•Р Р†Р С•
            result = await db.execute(
                select(User)
                .options(selectinload(User.roles).selectinload(UserRole.role))
                .where(User.id == current_user.id)
            )
            user_with_roles = result.scalar_one()
            user_roles = [ur.role.name for ur in user_with_roles.roles]
        
        logger.debug(f"User roles: {user_roles}")
        
        # Р С’Р Т‘Р СР С‘Р Р… Р С‘Р СР ВµР ВµРЎвЂљ Р Р†РЎРѓР Вµ Р С—РЎР‚Р В°Р Р†Р В° - Р С—РЎР‚Р С•Р С—РЎС“РЎРѓР С”Р В°Р ВµР С Р С—РЎР‚Р С•Р Р†Р ВµРЎР‚Р С”РЎС“
        if "admin" in user_roles:
            logger.debug("User is admin, skipping permission check")
        else:
            logger.debug("User is not admin, checking permissions")
            await require_permission(current_user, "equipment:create", db)
        
        logger.debug(f"Checking passport number: {equipment_data.passport_number}")
        # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚Р С”Р В° РЎС“Р Р…Р С‘Р С”Р В°Р В»РЎРЉР Р…Р С•РЎРѓРЎвЂљР С‘ Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљР В°
        result = await db.execute(
            select(Equipment).where(Equipment.passport_number == equipment_data.passport_number)
        )
        if result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Passport number already exists")
        
        if equipment_data.inventory_number:
            result = await db.execute(
                select(Equipment).where(Equipment.inventory_number == equipment_data.inventory_number)
            )
            if result.scalar_one_or_none():
                raise HTTPException(status_code=400, detail="Inventory number already exists")
        
        logger.debug("Creating equipment object")
        new_equipment = Equipment(
            **equipment_data.dict(),
            created_by=current_user.id,
        )
        db.add(new_equipment)
        logger.debug("Flushing to database")
        await db.flush()
        logger.info(f"Equipment created with ID: {new_equipment.id}")
        
        # Р вЂєР С•Р С–Р С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
        logger.debug("Creating activity log")
        activity = UserActivity(
            user_id=current_user.id,
            action_type="create",
            entity_type="equipment",
            entity_id=new_equipment.id,
            description=f"Created equipment {new_equipment.passport_number}"
        )
        db.add(activity)
        
        logger.debug("Committing to database")
        await db.commit()
        logger.debug("Refreshing equipment")
        await db.refresh(new_equipment)
        logger.info("Equipment saved successfully")
        
        return _equipment_to_response(new_equipment)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating equipment: {e}", exc_info=True)
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@router.put("/{equipment_id}", response_model=EquipmentResponse)
async def update_equipment(
    equipment_id: int,
    equipment_data: EquipmentUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р С›Р В±Р Р…Р С•Р Р†Р С‘РЎвЂљРЎРЉ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ"""
    await require_permission(current_user, "equipment:update", db)
    
    result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equipment = result.scalar_one_or_none()
    
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    
    # Р РЋР С•РЎвЂ¦РЎР‚Р В°Р Р…Р ВµР Р…Р С‘Р Вµ Р С‘РЎРѓРЎвЂљР С•РЎР‚Р С‘Р С‘ Р С‘Р В·Р СР ВµР Р…Р ВµР Р…Р С‘Р в„–
    update_data = equipment_data.dict(exclude_unset=True)
    for field, new_value in update_data.items():
        if hasattr(equipment, field):
            old_value = getattr(equipment, field)
            if old_value != new_value:
                history = EquipmentHistory(
                    equipment_id=equipment.id,
                    changed_by=current_user.id,
                    field_name=field,
                    old_value=str(old_value) if old_value else None,
                    new_value=str(new_value) if new_value else None
                )
                db.add(history)
                setattr(equipment, field, new_value)
    
    # Р вЂєР С•Р С–Р С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
    activity = UserActivity(
        user_id=current_user.id,
        action_type="update",
        entity_type="equipment",
        entity_id=equipment.id,
        description=f"Updated equipment {equipment.passport_number}"
    )
    db.add(activity)
    
    await db.commit()
    await db.refresh(equipment)
    
    return _equipment_to_response(equipment)


@router.post("/bulk", response_model=EquipmentBulkResponse)
async def bulk_create_equipment(
    payload: EquipmentBulkRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СљР В°РЎРѓРЎРѓР С•Р Р†Р С•Р Вµ Р Т‘Р С•Р В±Р В°Р Р†Р В»Р ВµР Р…Р С‘Р Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ"""
    await require_permission(current_user, "equipment:create", db)

    return await _bulk_create_equipment_items(
        items=payload.items,
        skip_duplicates=payload.skip_duplicates,
        current_user=current_user,
        db=db
    )


def _parse_equipment_csv_text(decoded: str) -> (List[EquipmentBulkItem], List[dict]):
    """
    Р С›Р В±РЎвЂ°Р С‘Р в„– CSV-Р С—Р В°РЎР‚РЎРѓР ВµРЎР‚ Р Т‘Р В»РЎРЏ Р СР В°РЎРѓРЎРѓР С•Р Р†Р С•Р С–Р С• Р С‘Р СР С—Р С•РЎР‚РЎвЂљР В° Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ.
    Р ВРЎРѓР С—Р С•Р В»РЎРЉР В·РЎС“Р ВµРЎвЂљРЎРѓРЎРЏ Р С”Р В°Р С” Р Т‘Р В»РЎРЏ Р В·Р В°Р С–РЎР‚РЎС“Р В·Р С”Р С‘ РЎвЂћР В°Р в„–Р В»Р С•Р Р†, РЎвЂљР В°Р С” Р С‘ Р Т‘Р В»РЎРЏ OCR-Р С‘Р СР С—Р С•РЎР‚РЎвЂљР В°.
    """
    # Р С›Р С—РЎР‚Р ВµР Т‘Р ВµР В»РЎРЏР ВµР С РЎР‚Р В°Р В·Р Т‘Р ВµР В»Р С‘РЎвЂљР ВµР В»РЎРЉ (Р С—Р С•Р Т‘Р Т‘Р ВµРЎР‚Р В¶Р С”Р В° Р С”Р В°Р С” Р В·Р В°Р С—РЎРЏРЎвЂљР С•Р в„–, РЎвЂљР В°Р С” Р С‘ РЎвЂљР С•РЎвЂЎР С”Р С‘ РЎРѓ Р В·Р В°Р С—РЎРЏРЎвЂљР С•Р в„–)
    first_line = decoded.splitlines()[0] if decoded.splitlines() else ""
    delimiter = ";" if first_line.count(";") > first_line.count(",") else ","
    reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV file is missing headers")

    parsed_items: List[EquipmentBulkItem] = []
    parse_errors: List[dict] = []
    required_fields = {"equipment_type", "passport_number"}
    headers = {h.strip().lower() for h in reader.fieldnames if h}
    if not required_fields.issubset(headers):
        raise HTTPException(
            status_code=400,
            detail="CSV file must contain at least 'equipment_type' and 'passport_number' columns",
        )

    # Р РЋР С—Р С‘РЎРѓР С•Р С” Р С”Р В»РЎР‹РЎвЂЎР ВµР Р†РЎвЂ№РЎвЂ¦ РЎРѓР В»Р С•Р Р†, Р С”Р С•РЎвЂљР С•РЎР‚РЎвЂ№Р Вµ РЎС“Р С”Р В°Р В·РЎвЂ№Р Р†Р В°РЎР‹РЎвЂљ Р Р…Р В° РЎРѓРЎвЂљРЎР‚Р С•Р С”РЎС“ РЎРѓ Р С—Р С•Р Т‘РЎРѓР С”Р В°Р В·Р С”Р В°Р СР С‘
    hint_keywords = [
        "Р С•Р В±РЎРЏР В·Р В°РЎвЂљР ВµР В»РЎРЉР Р…Р С•",
        "Р Р…Р ВµР С•Р В±РЎРЏР В·Р В°РЎвЂљР ВµР В»РЎРЉР Р…Р С•",
        "Р Р…Р В°Р С—РЎР‚Р С‘Р СР ВµРЎР‚",
        "РЎвЂљР С‘Р С— Р С—РЎРѓ",
        "Р Р…Р С•Р СР ВµРЎР‚ Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљР В°",
        "Р С‘Р Р…Р Р†Р ВµР Р…РЎвЂљР В°РЎР‚Р Р…РЎвЂ№Р в„– Р Р…Р С•Р СР ВµРЎР‚",
        "Р С—Р С•Р В·Р С‘РЎвЂ Р С‘РЎРЏ",
        "РЎвЂ Р ВµРЎвЂ¦",
        "Р С–РЎР‚РЎС“Р В·Р С•Р С—Р С•Р Т‘РЎР‰Р ВµР СР Р…Р С•РЎРѓРЎвЂљРЎРЉ",
        "Р В·Р В°Р Р†Р С•Р Т‘",
        "Р СР ВµРЎРѓРЎвЂљР С• РЎС“РЎРѓРЎвЂљР В°Р Р…Р С•Р Р†Р С”Р С‘",
        "Р Т‘Р В°РЎвЂљР В° Р Р†Р Р†Р С•Р Т‘Р В°",
        "Р Т‘Р В°РЎвЂљР В° Р С—РЎвЂљР С•",
        "Р Т‘Р В°РЎвЂљР В° РЎвЂЎРЎвЂљР С•",
        "РЎРѓРЎвЂљР В°РЎвЂљРЎС“РЎРѓ",
    ]

    for row_index, row in enumerate(reader, start=2):  # Р Р€РЎвЂЎР С‘РЎвЂљРЎвЂ№Р Р†Р В°Р ВµР С РЎРѓРЎвЂљРЎР‚Р С•Р С”РЎС“ Р В·Р В°Р С–Р С•Р В»Р С•Р Р†Р С”Р В°
        if not row:
            continue
        # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С, Р ВµРЎРѓРЎвЂљРЎРЉ Р В»Р С‘ Р Т‘Р В°Р Р…Р Р…РЎвЂ№Р Вµ Р Р† РЎРѓРЎвЂљРЎР‚Р С•Р С”Р Вµ
        if not any((value or "").strip() for value in row.values()):
            continue

        # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С, РЎРЏР Р†Р В»РЎРЏР ВµРЎвЂљРЎРѓРЎРЏ Р В»Р С‘ РЎРѓРЎвЂљРЎР‚Р С•Р С”Р В° Р С—Р С•Р Т‘РЎРѓР С”Р В°Р В·Р С”Р В°Р СР С‘ (РЎРѓР С•Р Т‘Р ВµРЎР‚Р В¶Р С‘РЎвЂљ Р С”Р В»РЎР‹РЎвЂЎР ВµР Р†РЎвЂ№Р Вµ РЎРѓР В»Р С•Р Р†Р В°)
        row_text = " ".join((value or "").lower() for value in row.values())
        if any(keyword in row_text for keyword in hint_keywords):
            continue  # Р СџРЎР‚Р С•Р С—РЎС“РЎРѓР С”Р В°Р ВµР С РЎРѓРЎвЂљРЎР‚Р С•Р С”РЎС“ РЎРѓ Р С—Р С•Р Т‘РЎРѓР С”Р В°Р В·Р С”Р В°Р СР С‘

        normalized = {
            "equipment_type": (row.get("equipment_type") or "").strip(),
            "passport_number": (row.get("passport_number") or "").strip(),
            "inventory_number": (row.get("inventory_number") or "").strip() or None,
            "position": (row.get("position") or "").strip() or None,
            "workshop": (row.get("workshop") or "").strip() or None,
            "load_capacity": _normalize_float(row.get("load_capacity")),
            "manufacturer": (row.get("manufacturer") or "").strip() or None,
            "installation_location": (row.get("installation_location") or "").strip() or None,
            "installation_date": _normalize_csv_date(row.get("installation_date")),
            "pto_date": _normalize_csv_date(row.get("pto_date")),
            "cto_date": _normalize_csv_date(row.get("cto_date")),
            "expertise_date": _normalize_csv_date(row.get("expertise_date")),
            "operation_permit_until": _normalize_csv_date(row.get("operation_permit_until")),
            "operation_banned": (
                _normalize_bool(row.get("operation_banned"))
                if row.get("operation_banned") is not None
                else _normalize_bool(row.get("ban_on_operation"))
            ),
            "epb_positive_details": (row.get("epb_positive_details") or "").strip() or None,
            "rostekhnadzor_registered": (
                _normalize_bool(row.get("rostekhnadzor_registered"))
                if row.get("rostekhnadzor_registered") is not None
                else _normalize_bool(row.get("registered_in_rostekhnadzor"))
            ),
            "status": (row.get("status") or "active").strip() or "active",
        }

        try:
            parsed_items.append(EquipmentBulkItem(**normalized))
        except ValidationError as exc:
            parse_errors.append(
                {
                    "row": row_index,
                    "detail": exc.errors(),
                }
            )

    return parsed_items, parse_errors


@router.post("/bulk/upload", response_model=EquipmentBulkResponse)
async def bulk_upload_equipment(
    file: UploadFile = File(),
    skip_duplicates: bool = True,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Р СљР В°РЎРѓРЎРѓР С•Р Р†Р С•Р Вµ Р Т‘Р С•Р В±Р В°Р Р†Р В»Р ВµР Р…Р С‘Р Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ РЎвЂЎР ВµРЎР‚Р ВµР В· CSV"""
    await require_permission(current_user, "equipment:create", db)

    filename = file.filename or ""
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a CSV file")

    raw_content = await file.read()
    if not raw_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        decoded = raw_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        decoded = raw_content.decode("utf-8")

    parsed_items, parse_errors = _parse_equipment_csv_text(decoded)

    if not parsed_items:
        raise HTTPException(status_code=400, detail="CSV file does not contain valid rows")

    response = await _bulk_create_equipment_items(
        items=parsed_items,
        skip_duplicates=skip_duplicates,
        current_user=current_user,
        db=db,
    )

    response.errors.extend(parse_errors)
    response.skipped += len(parse_errors)
    return response

@router.delete("/{equipment_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_equipment(
    equipment_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р Р€Р Т‘Р В°Р В»Р С‘РЎвЂљРЎРЉ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ"""
    await require_permission(current_user, "equipment:delete", db)
    
    result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equipment = result.scalar_one_or_none()
    
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    
    # Р вЂєР С•Р С–Р С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
    activity = UserActivity(
        user_id=current_user.id,
        action_type="delete",
        entity_type="equipment",
        entity_id=equipment.id,
        description=f"Deleted equipment {equipment.passport_number}"
    )
    db.add(activity)
    
    await db.delete(equipment)
    await db.commit()
    return None

@router.get("/{equipment_id}/history", response_model=List[dict])
async def get_equipment_history(
    equipment_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СџР С•Р В»РЎС“РЎвЂЎР С‘РЎвЂљРЎРЉ Р С‘РЎРѓРЎвЂљР С•РЎР‚Р С‘РЎР‹ Р С‘Р В·Р СР ВµР Р…Р ВµР Р…Р С‘Р в„– Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ"""
    await require_permission(current_user, "equipment:read", db)
    
    result = await db.execute(
        select(EquipmentHistory)
        .where(EquipmentHistory.equipment_id == equipment_id)
        .order_by(EquipmentHistory.created_at.desc())
    )
    history = result.scalars().all()
    
    return [
        {
            "id": h.id,
            "field_name": h.field_name,
            "old_value": h.old_value,
            "new_value": h.new_value,
            "changed_by": h.changed_by,
            "created_at": h.created_at.isoformat()
        }
        for h in history
    ]

@router.put("/bulk/update", response_model=EquipmentBulkUpdateResponse)
async def bulk_update_equipment(
    payload: EquipmentBulkUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СљР В°РЎРѓРЎРѓР С•Р Р†Р С•Р Вµ РЎР‚Р ВµР Т‘Р В°Р С”РЎвЂљР С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ"""
    await require_permission(current_user, "equipment:update", db)

    if not payload.equipment_ids:
        raise HTTPException(status_code=400, detail="Equipment IDs are required")

    updated = 0
    errors: List[dict] = []
    update_data = payload.update_data.dict(exclude_unset=True)

    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")

    for eq_id in payload.equipment_ids:
        try:
            result = await db.execute(select(Equipment).where(Equipment.id == eq_id))
            equipment = result.scalar_one_or_none()
            
            if not equipment:
                errors.append({"equipment_id": eq_id, "detail": "Equipment not found"})
                continue

            # Р РЋР С•РЎвЂ¦РЎР‚Р В°Р Р…Р ВµР Р…Р С‘Р Вµ Р С‘РЎРѓРЎвЂљР С•РЎР‚Р С‘Р С‘ Р С‘Р В·Р СР ВµР Р…Р ВµР Р…Р С‘Р в„–
            for field, new_value in update_data.items():
                if hasattr(equipment, field):
                    old_value = getattr(equipment, field)
                    if old_value != new_value:
                        history = EquipmentHistory(
                            equipment_id=equipment.id,
                            changed_by=current_user.id,
                            field_name=field,
                            old_value=str(old_value) if old_value else None,
                            new_value=str(new_value) if new_value else None
                        )
                        db.add(history)
                        setattr(equipment, field, new_value)

            # Р вЂєР С•Р С–Р С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
            activity = UserActivity(
                user_id=current_user.id,
                action_type="update",
                entity_type="equipment",
                entity_id=equipment.id,
                description=f"Bulk updated equipment {equipment.passport_number}"
            )
            db.add(activity)
            updated += 1

        except Exception as exc:
            errors.append({"equipment_id": eq_id, "detail": str(exc)})

    await db.commit()

    return EquipmentBulkUpdateResponse(updated=updated, errors=errors)

@router.put("/bulk/dates", response_model=EquipmentBulkDatesResponse)
async def bulk_update_dates(
    payload: EquipmentBulkDatesRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СљР В°РЎРѓРЎРѓР С•Р Р†Р С•Р Вµ Р Р…Р В°Р В·Р Р…Р В°РЎвЂЎР ВµР Р…Р С‘Р Вµ Р Т‘Р В°РЎвЂљ Р СџР СћР С›/Р В§Р СћР С›"""
    await require_permission(current_user, "equipment:update", db)

    if not payload.equipment_ids:
        raise HTTPException(status_code=400, detail="Equipment IDs are required")

    if not payload.pto_date and not payload.cto_date:
        raise HTTPException(status_code=400, detail="At least one date (PTO or CTO) must be provided")

    updated = 0
    errors: List[dict] = []

    for eq_id in payload.equipment_ids:
        try:
            result = await db.execute(select(Equipment).where(Equipment.id == eq_id))
            equipment = result.scalar_one_or_none()
            
            if not equipment:
                errors.append({"equipment_id": eq_id, "detail": "Equipment not found"})
                continue

            # Р РЋР С•РЎвЂ¦РЎР‚Р В°Р Р…Р ВµР Р…Р С‘Р Вµ Р С‘РЎРѓРЎвЂљР С•РЎР‚Р С‘Р С‘ Р С‘Р В·Р СР ВµР Р…Р ВµР Р…Р С‘Р в„–
            if payload.pto_date is not None:
                old_pto = equipment.pto_date
                if old_pto != payload.pto_date:
                    history = EquipmentHistory(
                        equipment_id=equipment.id,
                        changed_by=current_user.id,
                        field_name="pto_date",
                        old_value=str(old_pto) if old_pto else None,
                        new_value=str(payload.pto_date) if payload.pto_date else None
                    )
                    db.add(history)
                    equipment.pto_date = payload.pto_date

            if payload.cto_date is not None:
                old_cto = equipment.cto_date
                if old_cto != payload.cto_date:
                    history = EquipmentHistory(
                        equipment_id=equipment.id,
                        changed_by=current_user.id,
                        field_name="cto_date",
                        old_value=str(old_cto) if old_cto else None,
                        new_value=str(payload.cto_date) if payload.cto_date else None
                    )
                    db.add(history)
                    equipment.cto_date = payload.cto_date

            # Р вЂєР С•Р С–Р С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
            activity = UserActivity(
                user_id=current_user.id,
                action_type="update",
                entity_type="equipment",
                entity_id=equipment.id,
                description=f"Bulk updated dates for equipment {equipment.passport_number}"
            )
            db.add(activity)
            updated += 1

        except Exception as exc:
            errors.append({"equipment_id": eq_id, "detail": str(exc)})

    await db.commit()

    return EquipmentBulkDatesResponse(updated=updated, errors=errors)

@router.post("/ocr-upsert", response_model=EquipmentOCRUpsertResponse)
async def ocr_upsert_equipment(
    equipment_data: EquipmentOCRUpsertRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р РЋР С•Р В·Р Т‘Р В°РЎвЂљРЎРЉ Р С‘Р В»Р С‘ Р С•Р В±Р Р…Р С•Р Р†Р С‘РЎвЂљРЎРЉ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ РЎвЂЎР ВµРЎР‚Р ВµР В· OCR Р Т‘Р В°Р Р…Р Р…РЎвЂ№Р Вµ"""
    await require_permission(current_user, "equipment:create", db)
    
    # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С Р С•Р В±РЎРЏР В·Р В°РЎвЂљР ВµР В»РЎРЉР Р…РЎвЂ№Р Вµ Р С—Р С•Р В»РЎРЏ
    if not equipment_data.passport_number and not equipment_data.inventory_number:
        raise HTTPException(
            status_code=400, 
            detail="Either passport_number or inventory_number is required"
        )
    
    # Р ВРЎвЂ°Р ВµР С РЎРѓРЎС“РЎвЂ°Р ВµРЎРѓРЎвЂљР Р†РЎС“РЎР‹РЎвЂ°Р ВµР Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
    existing_equipment = None
    
    # Р РЋР Р…Р В°РЎвЂЎР В°Р В»Р В° Р С‘РЎвЂ°Р ВµР С Р С—Р С• Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљР Р…Р С•Р СРЎС“ Р Р…Р С•Р СР ВµРЎР‚РЎС“
    if equipment_data.passport_number:
        result = await db.execute(
            select(Equipment).where(Equipment.passport_number == equipment_data.passport_number)
        )
        existing_equipment = result.scalar_one_or_none()
    
    # Р вЂўРЎРѓР В»Р С‘ Р Р…Р Вµ Р Р…Р В°Р в„–Р Т‘Р ВµР Р…Р С• Р С—Р С• Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљРЎС“, Р С‘РЎвЂ°Р ВµР С Р С—Р С• Р С‘Р Р…Р Р†Р ВµР Р…РЎвЂљР В°РЎР‚Р Р…Р С•Р СРЎС“ Р Р…Р С•Р СР ВµРЎР‚РЎС“
    if not existing_equipment and equipment_data.inventory_number:
        result = await db.execute(
            select(Equipment).where(Equipment.inventory_number == equipment_data.inventory_number)
        )
        existing_equipment = result.scalar_one_or_none()
    
    # Р вЂўРЎРѓР В»Р С‘ Р Р…Р Вµ Р Р…Р В°Р в„–Р Т‘Р ВµР Р…Р С• Р С—Р С• Р С‘Р Р…Р Р†Р ВµР Р…РЎвЂљР В°РЎР‚Р Р…Р С•Р СРЎС“, Р С‘РЎвЂ°Р ВµР С Р С—Р С• Р С‘Р СР ВµР Р…Р С‘ (Р ВµРЎРѓР В»Р С‘ РЎС“Р С”Р В°Р В·Р В°Р Р…Р С•)
    if not existing_equipment and equipment_data.name:
        # Р ВРЎвЂ°Р ВµР С Р С—Р С• Р С”Р С•Р СР В±Р С‘Р Р…Р В°РЎвЂ Р С‘Р С‘ РЎвЂљР С‘Р С—Р В° Р С‘ Р С—Р С•Р В·Р С‘РЎвЂ Р С‘Р С‘
        search_conditions = []
        if equipment_data.equipment_type:
            search_conditions.append(Equipment.equipment_type.ilike(f"%{equipment_data.equipment_type}%"))
        if equipment_data.position:
            search_conditions.append(Equipment.position.ilike(f"%{equipment_data.position}%"))
        
        if search_conditions:
            result = await db.execute(
                select(Equipment).where(and_(*search_conditions))
            )
            potential_matches = result.scalars().all()
            
            # Р вЂўРЎРѓР В»Р С‘ Р Р…Р В°Р в„–Р Т‘Р ВµР Р…Р С• РЎвЂљР С•Р В»РЎРЉР С”Р С• Р С•Р Т‘Р Р…Р С• РЎРѓР С•Р Р†Р С—Р В°Р Т‘Р ВµР Р…Р С‘Р Вµ, Р С‘РЎРѓР С—Р С•Р В»РЎРЉР В·РЎС“Р ВµР С Р ВµР С–Р С•
            if len(potential_matches) == 1:
                existing_equipment = potential_matches[0]
    
    created = False
    
    if existing_equipment:
        # Р С›Р В±Р Р…Р С•Р Р†Р В»РЎРЏР ВµР С РЎРѓРЎС“РЎвЂ°Р ВµРЎРѓРЎвЂљР Р†РЎС“РЎР‹РЎвЂ°Р ВµР Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
        logger.info(f"Updating existing equipment ID {existing_equipment.id}")
        
        # Р С›Р В±Р Р…Р С•Р Р†Р В»РЎРЏР ВµР С РЎвЂљР С•Р В»РЎРЉР С”Р С• Р Р…Р ВµР С—РЎС“РЎРѓРЎвЂљРЎвЂ№Р Вµ Р С—Р С•Р В»РЎРЏ
        update_fields = {}
        
        if equipment_data.name and not existing_equipment.equipment_type:
            update_fields['equipment_type'] = equipment_data.name
        elif equipment_data.equipment_type:
            update_fields['equipment_type'] = equipment_data.equipment_type
            
        if equipment_data.capacity and not existing_equipment.load_capacity:
            update_fields['load_capacity'] = equipment_data.capacity
            
        if equipment_data.inventory_number and not existing_equipment.inventory_number:
            update_fields['inventory_number'] = equipment_data.inventory_number
            
        if equipment_data.manufacturer and not existing_equipment.manufacturer:
            update_fields['manufacturer'] = equipment_data.manufacturer
            
        if equipment_data.installation_date and not existing_equipment.installation_date:
            update_fields['installation_date'] = equipment_data.installation_date
            
        if equipment_data.pto_date and not existing_equipment.pto_date:
            update_fields['pto_date'] = equipment_data.pto_date
            
        if equipment_data.cto_date and not existing_equipment.cto_date:
            update_fields['cto_date'] = equipment_data.cto_date
            
        if equipment_data.position and not existing_equipment.position:
            update_fields['position'] = equipment_data.position
            
        if equipment_data.workshop and not existing_equipment.workshop:
            update_fields['workshop'] = equipment_data.workshop
        
        # Р СџРЎР‚Р С‘Р СР ВµР Р…РЎРЏР ВµР С Р С•Р В±Р Р…Р С•Р Р†Р В»Р ВµР Р…Р С‘РЎРЏ
        for field, value in update_fields.items():
            setattr(existing_equipment, field, value)
        
        # Р РЋР С•РЎвЂ¦РЎР‚Р В°Р Р…РЎРЏР ВµР С Р С‘РЎРѓРЎвЂљР С•РЎР‚Р С‘РЎР‹ Р С‘Р В·Р СР ВµР Р…Р ВµР Р…Р С‘Р в„– Р Т‘Р В»РЎРЏ Р С•Р В±Р Р…Р С•Р Р†Р В»Р ВµР Р…Р Р…РЎвЂ№РЎвЂ¦ Р С—Р С•Р В»Р ВµР в„–
        for field, new_value in update_fields.items():
            history = EquipmentHistory(
                equipment_id=existing_equipment.id,
                changed_by=current_user.id,
                field_name=field,
                old_value=None,  # Р вЂРЎвЂ№Р В»Р С• Р С—РЎС“РЎРѓРЎвЂљР С•Р Вµ
                new_value=str(new_value) if new_value else None
            )
            db.add(history)
        
        equipment_id = existing_equipment.id
        
    else:
        # Р РЋР С•Р В·Р Т‘Р В°Р ВµР С Р Р…Р С•Р Р†Р С•Р Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
        logger.info("Creating new equipment from OCR data")
        
        # Р С›Р С—РЎР‚Р ВµР Т‘Р ВµР В»РЎРЏР ВµР С Р С•Р В±РЎРЏР В·Р В°РЎвЂљР ВµР В»РЎРЉР Р…РЎвЂ№Р Вµ Р С—Р С•Р В»РЎРЏ
        equipment_type = equipment_data.equipment_type or equipment_data.name or "Р СњР ВµР С•Р С—РЎР‚Р ВµР Т‘Р ВµР В»Р ВµР Р…Р Р…РЎвЂ№Р в„– РЎвЂљР С‘Р С—"
        passport_number = equipment_data.passport_number
        
        # Р вЂўРЎРѓР В»Р С‘ Р Р…Р ВµРЎвЂљ Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљР Р…Р С•Р С–Р С• Р Р…Р С•Р СР ВµРЎР‚Р В°, Р С–Р ВµР Р…Р ВµРЎР‚Р С‘РЎР‚РЎС“Р ВµР С Р Р†РЎР‚Р ВµР СР ВµР Р…Р Р…РЎвЂ№Р в„–
        if not passport_number:
            import random
            passport_number = f"OCR-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        
        # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С РЎС“Р Р…Р С‘Р С”Р В°Р В»РЎРЉР Р…Р С•РЎРѓРЎвЂљРЎРЉ Р С—Р В°РЎРѓР С—Р С•РЎР‚РЎвЂљР Р…Р С•Р С–Р С• Р Р…Р С•Р СР ВµРЎР‚Р В°
        result = await db.execute(
            select(Equipment).where(Equipment.passport_number == passport_number)
        )
        if result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Passport number already exists")
        
        # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С РЎС“Р Р…Р С‘Р С”Р В°Р В»РЎРЉР Р…Р С•РЎРѓРЎвЂљРЎРЉ Р С‘Р Р…Р Р†Р ВµР Р…РЎвЂљР В°РЎР‚Р Р…Р С•Р С–Р С• Р Р…Р С•Р СР ВµРЎР‚Р В° Р ВµРЎРѓР В»Р С‘ РЎС“Р С”Р В°Р В·Р В°Р Р…
        if equipment_data.inventory_number:
            result = await db.execute(
                select(Equipment).where(Equipment.inventory_number == equipment_data.inventory_number)
            )
            if result.scalar_one_or_none():
                raise HTTPException(status_code=400, detail="Inventory number already exists")
        
        new_equipment = Equipment(
            equipment_type=equipment_type,
            passport_number=passport_number,
            inventory_number=equipment_data.inventory_number,
            load_capacity=equipment_data.capacity,
            manufacturer=equipment_data.manufacturer,
            installation_date=equipment_data.installation_date,
            pto_date=equipment_data.pto_date,
            cto_date=equipment_data.cto_date,
            position=equipment_data.position,
            workshop=equipment_data.workshop,
            status="active",
            created_by=current_user.id
        )
        db.add(new_equipment)
        await db.flush()
        
        equipment_id = new_equipment.id
        created = True
    
    # Р вЂєР С•Р С–Р С‘РЎР‚Р С•Р Р†Р В°Р Р…Р С‘Р Вµ
    action = "create" if created else "update"
    activity = UserActivity(
        user_id=current_user.id,
        action_type=action,
        entity_type="equipment",
        entity_id=equipment_id,
        description=f"OCR {action}d equipment: {equipment_data.passport_number or equipment_data.inventory_number}"
    )
    db.add(activity)
    
    await db.commit()
    
    logger.info(f"OCR upsert completed: equipment_id={equipment_id}, created={created}")
    
    return EquipmentOCRUpsertResponse(
        id=equipment_id,
        created=created
    )


@router.post("/ocr-import", response_model=EquipmentBulkResponse)
async def ocr_import_equipment(
    payload: EquipmentOCRImportRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Р СљР В°РЎРѓРЎРѓР С•Р Р†РЎвЂ№Р в„– Р С‘Р СР С—Р С•РЎР‚РЎвЂљ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ Р Р…Р В° Р С•РЎРѓР Р…Р С•Р Р†Р Вµ OCR-РЎР‚Р ВµР В·РЎС“Р В»РЎРЉРЎвЂљР В°РЎвЂљР В°.

    Р вЂ™Р В°РЎР‚Р С‘Р В°Р Р…РЎвЂљРЎвЂ№ Р С‘РЎРѓР С—Р С•Р В»РЎРЉР В·Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ:
    - Р вЂР С•РЎвЂљ Р С‘Р В»Р С‘ Р Р†Р Р…Р ВµРЎв‚¬Р Р…Р С‘Р в„– РЎРѓР ВµРЎР‚Р Р†Р С‘РЎРѓ РЎР‚Р В°РЎРѓР С—Р С•Р В·Р Р…Р В°Р В» РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ РЎС“ Р С‘ Р С—РЎР‚Р С‘РЎРѓР В»Р В°Р В» CSV-РЎвЂљР ВµР С”РЎРѓРЎвЂљ Р Р† ocr_text
    - Р Р€Р С”Р В°Р В·Р В°Р Р… file_id Р Р…Р В° РЎР‚Р В°Р Р…Р ВµР Вµ Р В·Р В°Р С–РЎР‚РЎС“Р В¶Р ВµР Р…Р Р…РЎвЂ№Р в„– CSV-РЎвЂћР В°Р в„–Р В» Р Р† РЎвЂљР В°Р В±Р В»Р С‘РЎвЂ Р Вµ files

    Р СџР С•Р Т‘Р Т‘Р ВµРЎР‚Р В¶Р С”Р В° РЎР‚Р В°РЎРѓР С—Р С•Р В·Р Р…Р В°Р Р†Р В°Р Р…Р С‘РЎРЏ РЎвЂћР С•РЎвЂљР С• (image РІвЂ вЂ™ text) Р Т‘Р С•Р В»Р В¶Р Р…Р В° Р В±РЎвЂ№РЎвЂљРЎРЉ РЎР‚Р ВµР В°Р В»Р С‘Р В·Р С•Р Р†Р В°Р Р…Р В°
    Р Р†Р С• Р Р†Р Р…Р ВµРЎв‚¬Р Р…Р ВµР С РЎРѓР ВµРЎР‚Р Р†Р С‘РЎРѓР Вµ, Р С”Р С•РЎвЂљР С•РЎР‚РЎвЂ№Р в„– Р С—Р ВµРЎР‚Р ВµР Т‘Р В°РЎРѓРЎвЂљ РЎС“Р В¶Р Вµ Р С–Р С•РЎвЂљР С•Р Р†РЎвЂ№Р в„– РЎвЂљР В°Р В±Р В»Р С‘РЎвЂЎР Р…РЎвЂ№Р в„– РЎвЂљР ВµР С”РЎРѓРЎвЂљ.
    """
    await require_permission(current_user, "equipment:create", db)

    if not payload.ocr_text and not payload.file_id:
        raise HTTPException(
            status_code=400,
            detail="Either ocr_text or file_id must be provided",
        )

    decoded = None

    # Р вЂўРЎРѓР В»Р С‘ Р С—РЎР‚Р С‘РЎв‚¬РЎвЂР В» Р С–Р С•РЎвЂљР С•Р Р†РЎвЂ№Р в„– РЎвЂљР ВµР С”РЎРѓРЎвЂљ (Р Р…Р В°Р С—РЎР‚Р С‘Р СР ВµРЎР‚, Р С•РЎвЂљ Telegram-Р В±Р С•РЎвЂљР В° Р С—Р С•РЎРѓР В»Р Вµ OCR)
    if payload.ocr_text:
        decoded = payload.ocr_text

    # Р вЂўРЎРѓР В»Р С‘ РЎС“Р С”Р В°Р В·Р В°Р Р… file_id РІР‚вЂќ Р С—РЎР‚Р С•Р В±РЎС“Р ВµР С Р С—РЎР‚Р С•РЎвЂЎР С‘РЎвЂљР В°РЎвЂљРЎРЉ РЎРѓР С•Р Т‘Р ВµРЎР‚Р В¶Р С‘Р СР С•Р Вµ РЎвЂћР В°Р в„–Р В»Р В°
    if not decoded and payload.file_id:
        result = await db.execute(select(File).where(File.id == payload.file_id))
        file_obj: Optional[File] = result.scalar_one_or_none()
        if not file_obj:
            raise HTTPException(status_code=404, detail="File not found")

        if not file_obj.file_path:
            raise HTTPException(
                status_code=400,
                detail="File has no file_path, cannot read from disk",
            )

        # Р С›Р С—РЎР‚Р ВµР Т‘Р ВµР В»РЎРЏР ВµР С, РЎРЏР Р†Р В»РЎРЏР ВµРЎвЂљРЎРѓРЎРЏ Р В»Р С‘ РЎвЂћР В°Р в„–Р В» РЎвЂљР ВµР С”РЎРѓРЎвЂљР С•Р С/CSV Р С‘Р В»Р С‘ Р С”Р В°РЎР‚РЎвЂљР С‘Р Р…Р С”Р С•Р в„–
        file_path = Path(file_obj.file_path)
        suffix = file_path.suffix.lower()

        # Р вЂўРЎРѓР В»Р С‘ РЎРЊРЎвЂљР С• Р С‘Р В·Р С•Р В±РЎР‚Р В°Р В¶Р ВµР Р…Р С‘Р Вµ РІР‚вЂќ Р В·Р В°Р С—РЎС“РЎРѓР С”Р В°Р ВµР С OCR
        if suffix in {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"} or (
            file_obj.mime_type and file_obj.mime_type.startswith("image/")
        ):
            # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С РЎРѓРЎС“РЎвЂ°Р ВµРЎРѓРЎвЂљР Р†Р С•Р Р†Р В°Р Р…Р С‘Р Вµ РЎвЂћР В°Р в„–Р В»Р В° Р С‘ Р Р…Р С•РЎР‚Р СР В°Р В»Р С‘Р В·РЎС“Р ВµР С Р С—РЎС“РЎвЂљРЎРЉ
            # Р В¤Р В°Р в„–Р В»РЎвЂ№ РЎРѓР С•РЎвЂ¦РЎР‚Р В°Р Р…РЎРЏРЎР‹РЎвЂљРЎРѓРЎРЏ Р С”Р В°Р С” Р С•РЎвЂљР Р…Р С•РЎРѓР С‘РЎвЂљР ВµР В»РЎРЉР Р…РЎвЂ№Р Вµ Р С—РЎС“РЎвЂљР С‘ РЎвЂљР С‘Р С—Р В° "uploads/filename"
            file_path_str = file_obj.file_path
            
            # Р СџРЎР‚Р С•Р В±РЎС“Р ВµР С РЎР‚Р В°Р В·Р Р…РЎвЂ№Р Вµ Р Р†Р В°РЎР‚Р С‘Р В°Р Р…РЎвЂљРЎвЂ№ Р С—РЎС“РЎвЂљР С‘
            possible_paths = [
                file_path_str,  # Р С™Р В°Р С” Р ВµРЎРѓРЎвЂљРЎРЉ (Р ВµРЎРѓР В»Р С‘ Р В°Р В±РЎРѓР С•Р В»РЎР‹РЎвЂљР Р…РЎвЂ№Р в„–)
                os.path.join("/app/backend", file_path_str),  # Р С›РЎвЂљР Р…Р С•РЎРѓР С‘РЎвЂљР ВµР В»РЎРЉР Р…Р С• backend
                os.path.join("/app", file_path_str),  # Р С›РЎвЂљР Р…Р С•РЎРѓР С‘РЎвЂљР ВµР В»РЎРЉР Р…Р С• Р С”Р С•РЎР‚Р Р…РЎРЏ
            ]
            
            actual_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    actual_path = path
                    break
            
            if not actual_path:
                # Р вЂўРЎРѓР В»Р С‘ Р Р…Р С‘ Р С•Р Т‘Р С‘Р Р… Р С—РЎС“РЎвЂљРЎРЉ Р Р…Р Вµ Р Р…Р В°Р в„–Р Т‘Р ВµР Р…, Р С—РЎР‚Р С•Р В±РЎС“Р ВµР С Р Р…Р В°Р в„–РЎвЂљР С‘ РЎвЂћР В°Р в„–Р В» Р С—Р С• Р С‘Р СР ВµР Р…Р С‘ Р Р† uploads
                filename = os.path.basename(file_path_str)
                uploads_path = os.path.join("/app/backend", "uploads", filename)
                if os.path.exists(uploads_path):
                    actual_path = uploads_path
                else:
                    raise HTTPException(
                        status_code=404,
                        detail=f"Image file not found on disk. Tried: {possible_paths + [uploads_path]}",
                    )
            
            logger.info(f"Running OCR for equipment import on image file: {actual_path}")
            decoded = _ocr_image_to_text(actual_path)
        else:
            # Р ВР Р…Р В°РЎвЂЎР Вµ РЎРѓРЎвЂЎР С‘РЎвЂљР В°Р ВµР С, РЎвЂЎРЎвЂљР С• РЎРЊРЎвЂљР С• РЎвЂљР ВµР С”РЎРѓРЎвЂљ/CSV
            try:
                with open(file_obj.file_path, "rb") as f:
                    raw_content = f.read()
            except FileNotFoundError:
                raise HTTPException(
                    status_code=404,
                    detail=f"File not found on disk: {file_obj.file_path}",
                )
            except Exception as e:
                logger.error(f"Error reading file {file_obj.file_path}: {e}", exc_info=True)
                raise HTTPException(
                    status_code=500,
                    detail=f"Error reading file: {str(e)}",
                )

            try:
                decoded = raw_content.decode("utf-8-sig")
            except UnicodeDecodeError:
                decoded = raw_content.decode("utf-8", errors="ignore")

    if not decoded or not decoded.strip():
        raise HTTPException(
            status_code=400,
            detail="OCR text is empty or could not be decoded",
        )

    parsed_items, parse_errors = _parse_equipment_csv_text(decoded)

    if not parsed_items:
        raise HTTPException(
            status_code=400,
            detail="OCR text/CSV does not contain valid equipment rows",
        )

    response = await _bulk_create_equipment_items(
        items=parsed_items,
        skip_duplicates=True,
        current_user=current_user,
        db=db,
    )

    response.errors.extend(parse_errors)
    response.skipped += len(parse_errors)
    return response

@router.get("/{equipment_id}/violations")
async def get_equipment_violations(
    equipment_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status: Optional[str] = None,
    severity: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р СџР С•Р В»РЎС“РЎвЂЎР С‘РЎвЂљРЎРЉ Р Р…Р В°РЎР‚РЎС“РЎв‚¬Р ВµР Р…Р С‘РЎРЏ Р Т‘Р В»РЎРЏ Р С”Р С•Р Р…Р С”РЎР‚Р ВµРЎвЂљР Р…Р С•Р С–Р С• Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ"""
    await require_permission(current_user, "violations:read", db)
    
    # Р СџРЎР‚Р С•Р Р†Р ВµРЎР‚РЎРЏР ВµР С РЎРѓРЎС“РЎвЂ°Р ВµРЎРѓРЎвЂљР Р†Р С•Р Р†Р В°Р Р…Р С‘Р Вµ Р С•Р В±Р С•РЎР‚РЎС“Р Т‘Р С•Р Р†Р В°Р Р…Р С‘РЎРЏ
    eq_result = await db.execute(select(Equipment).where(Equipment.id == equipment_id))
    equipment = eq_result.scalar_one_or_none()
    
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    
    # Р ВР СР С—Р С•РЎР‚РЎвЂљР С‘РЎР‚РЎС“Р ВµР С Р СР С•Р Т‘Р ВµР В»РЎРЉ Violation Р ВµРЎРѓР В»Р С‘ Р ВµРЎвЂ°Р Вµ Р Р…Р Вµ Р С‘Р СР С—Р С•РЎР‚РЎвЂљР С‘РЎР‚Р С•Р Р†Р В°Р Р…Р В°
    try:
        from backend.models import Violation
    except ImportError:
        from ..models import Violation
    
    # Р РЋРЎвЂљРЎР‚Р С•Р С‘Р С Р В·Р В°Р С—РЎР‚Р С•РЎРѓ Р Р…Р В°РЎР‚РЎС“РЎв‚¬Р ВµР Р…Р С‘Р в„–
    query = select(Violation).where(Violation.equipment_id == equipment_id)
    
    if status:
        query = query.where(Violation.status == status)
    
    if severity:
        query = query.where(Violation.severity == severity)
    
    query = query.order_by(Violation.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    violations = result.scalars().all()
    
    # Р В¤Р С•РЎР‚Р СР С‘РЎР‚РЎС“Р ВµР С Р С•РЎвЂљР Р†Р ВµРЎвЂљ
    violations_data = []
    for violation in violations:
        violations_data.append({
            "id": violation.id,
            "description": violation.description,
            "fnp_clause": violation.fnp_clause,
            "gost_clause": violation.gost_clause,
            "severity": violation.severity,
            "location": violation.location,
            "deadline": violation.deadline.isoformat() if violation.deadline else None,
            "status": violation.status,
            "resolved_at": violation.resolved_at.isoformat() if violation.resolved_at else None,
            "created_at": violation.created_at.isoformat()
        })
    
    return {
        "items": violations_data,
        "equipment": {
            "id": equipment.id,
            "equipment_type": equipment.equipment_type,
            "passport_number": equipment.passport_number,
            "position": equipment.position,
            "workshop": equipment.workshop
        }
    }



