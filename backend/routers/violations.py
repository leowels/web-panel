from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, status, Query, Request
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from sqlalchemy.orm import selectinload
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
import re
from pydantic import BaseModel
import os
import csv
import io

# РџРѕРґРґРµСЂР¶РєР° Р·Р°РїСѓСЃРєР° РєР°Рє СЃРєСЂРёРїС‚Р° Рё РєР°Рє РјРѕРґСѓР»СЏ
try:
    from backend.models import Violation, Inspection, Equipment, UserActivity, User, ViolationSLARule, AuditLog, SystemSettings
    from backend.database import get_db
    from backend.auth import get_current_user, require_permission
    from backend.audit import log_audit_event, build_field_changes
except ImportError:
    from ..models import Violation, Inspection, Equipment, UserActivity, User, ViolationSLARule, AuditLog, SystemSettings
    from ..database import get_db
    from ..auth import get_current_user, require_permission
    from ..audit import log_audit_event, build_field_changes

router = APIRouter(prefix="/api/violations", tags=["violations"])

def _default_sla_days(severity: Optional[str]) -> int:
    mapping = {
        "critical": 7,
        "high": 15,
        "medium": 30,
        "low": 60,
    }
    if not severity:
        return 30
    return mapping.get(severity, 30)

async def _resolve_sla_rule(
    db: AsyncSession,
    violation_type: Optional[str],
    severity: Optional[str]
) -> Optional[ViolationSLARule]:
    result = await db.execute(select(ViolationSLARule).where(ViolationSLARule.is_active == True))
    rules = result.scalars().all()
    if not rules:
        return None
    best = None
    best_score = -1
    best_priority = 999999
    for rule in rules:
        if rule.violation_type and violation_type:
            if rule.violation_type.strip().lower() != violation_type.strip().lower():
                continue
        elif rule.violation_type and not violation_type:
            continue
        if rule.severity and severity:
            if rule.severity != severity:
                continue
        elif rule.severity and not severity:
            continue
        score = 0
        if rule.violation_type and violation_type:
            score += 2
        if rule.severity and severity:
            score += 1
        priority = rule.priority if rule.priority is not None else 100
        if score > best_score or (score == best_score and priority < best_priority):
            best = rule
            best_score = score
            best_priority = priority
    return best

async def _apply_sla_deadline(db: AsyncSession, violation: Violation) -> None:
    if violation.deadline:
        return
    base_time = violation.created_at or datetime.utcnow()
    rule = await _resolve_sla_rule(db, violation.violation_type, violation.severity)
    if rule:
        violation.deadline = base_time + timedelta(days=rule.days)
        violation.deadline_source = "sla"
        violation.deadline_rule_id = rule.id
        return
    days = _default_sla_days(violation.severity)
    violation.deadline = base_time + timedelta(days=days)
    violation.deadline_source = "sla_default"
    violation.deadline_rule_id = None

async def _update_overdue_flags(db: AsyncSession, violations: List[Violation]) -> None:
    now = datetime.utcnow()
    changed = False
    for violation in violations:
        should_overdue = (
            violation.status != "resolved"
            and violation.deadline is not None
            and violation.deadline < now
        )
        if violation.is_overdue != should_overdue:
            violation.is_overdue = should_overdue
            violation.overdue_at = now if should_overdue else None
            changed = True
    if changed:
        await db.commit()

def _equipment_summary(equipment: Optional[Equipment]) -> Optional[EquipmentSummary]:
    if not equipment:
        return None
    return EquipmentSummary(
        id=equipment.id,
        equipment_type=equipment.equipment_type,
        passport_number=equipment.passport_number,
        position=equipment.position,
        inventory_number=equipment.inventory_number,
        workshop=equipment.workshop,
    )


def _violation_to_response(violation: Violation) -> ViolationResponse:
    return ViolationResponse(
        id=violation.id,
        inspection_id=violation.inspection_id,
        equipment_id=violation.equipment_id,
        description=violation.description,
        fnp_clause=violation.fnp_clause,
        gost_clause=violation.gost_clause,
        severity=violation.severity,
        criticality_level=violation.criticality_level,
        violation_type=violation.violation_type,
        violation_type_description=violation.violation_type_description,
        norm_reference=violation.norm_reference,
        recommended_act_text=violation.recommended_act_text,
        requirements=violation.requirements if isinstance(violation.requirements, list) else None,
        source=violation.source,
        reported_by=violation.reported_by,
        attachment_meta=violation.attachment_meta,
        ai_classification=violation.ai_classification,
        ai_recommendations=violation.ai_recommendations,
        ai_payload_raw=violation.ai_payload_raw,
        location=violation.location,
        deadline=violation.deadline,
        deadline_source=violation.deadline_source,
        deadline_rule_id=violation.deadline_rule_id,
        is_overdue=violation.is_overdue,
        overdue_at=violation.overdue_at,
        status=violation.status,
        resolved_at=violation.resolved_at,
        created_at=violation.created_at,
        updated_at=violation.updated_at,
        equipment=_equipment_summary(getattr(violation, "equipment", None)),
    )

class ViolationCreate(BaseModel):
    inspection_id: Optional[int] = None
    equipment_id: int
    description: str
    fnp_clause: Optional[str] = None
    gost_clause: Optional[str] = None
    severity: str = "medium"  # low, medium, high, critical
    criticality_level: Optional[str] = None
    violation_type: Optional[str] = None
    violation_type_description: Optional[str] = None
    norm_reference: Optional[str] = None
    recommended_act_text: Optional[str] = None
    requirements: Optional[List[str]] = None
    source: Optional[str] = None
    reported_by: Optional[int] = None
    attachment_meta: Optional[Dict[str, Any]] = None
    ai_classification: Optional[Dict[str, Any]] = None
    ai_recommendations: Optional[Dict[str, Any]] = None
    ai_payload_raw: Optional[Dict[str, Any]] = None
    location: Optional[str] = None
    deadline: Optional[datetime] = None

class ViolationUpdate(BaseModel):
    description: Optional[str] = None
    fnp_clause: Optional[str] = None
    gost_clause: Optional[str] = None
    severity: Optional[str] = None
    criticality_level: Optional[str] = None
    violation_type: Optional[str] = None
    violation_type_description: Optional[str] = None
    norm_reference: Optional[str] = None
    recommended_act_text: Optional[str] = None
    requirements: Optional[List[str]] = None
    source: Optional[str] = None
    reported_by: Optional[int] = None
    attachment_meta: Optional[Dict[str, Any]] = None
    ai_classification: Optional[Dict[str, Any]] = None
    ai_recommendations: Optional[Dict[str, Any]] = None
    ai_payload_raw: Optional[Dict[str, Any]] = None
    location: Optional[str] = None
    deadline: Optional[datetime] = None
    status: Optional[str] = None

class EquipmentSummary(BaseModel):
    id: int
    equipment_type: str
    passport_number: str
    position: Optional[str]
    inventory_number: Optional[str]
    workshop: Optional[str]

    class Config:
        from_attributes = True

class ViolationResponse(BaseModel):
    id: int
    inspection_id: Optional[int]
    equipment_id: int
    description: str
    fnp_clause: Optional[str]
    gost_clause: Optional[str]
    severity: str
    criticality_level: Optional[str]
    violation_type: Optional[str]
    violation_type_description: Optional[str]
    norm_reference: Optional[str]
    recommended_act_text: Optional[str]
    requirements: Optional[List[str]]
    source: Optional[str]
    reported_by: Optional[int]
    attachment_meta: Optional[Dict[str, Any]]
    ai_classification: Optional[Dict[str, Any]]
    ai_recommendations: Optional[Dict[str, Any]]
    ai_payload_raw: Optional[Dict[str, Any]]
    location: Optional[str]
    deadline: Optional[datetime]
    deadline_source: Optional[str] = None
    deadline_rule_id: Optional[int] = None
    is_overdue: Optional[bool] = None
    overdue_at: Optional[datetime] = None
    status: str
    resolved_at: Optional[datetime]
    created_at: datetime
    updated_at: datetime
    equipment: Optional[EquipmentSummary] = None

    class Config:
        from_attributes = True

class AIGenerateViolationResponse(BaseModel):
    """РћС‚РІРµС‚ СЃ РёРЅС„РѕСЂРјР°С†РёРµР№ Рѕ СЃРіРµРЅРµСЂРёСЂРѕРІР°РЅРЅРѕРј РЅР°СЂСѓС€РµРЅРёРё Рё РёСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹С… РґРѕРєСѓРјРµРЅС‚Р°С…"""
    violation: ViolationResponse
    used_documents: List[dict] = []  # РЎРїРёСЃРѕРє РґРѕРєСѓРјРµРЅС‚РѕРІ РёР· Р±Р°Р·С‹ Р·РЅР°РЅРёР№, РєРѕС‚РѕСЂС‹Рµ Р±С‹Р»Рё РёСЃРїРѕР»СЊР·РѕРІР°РЅС‹

class AIGenerateViolationRequest(BaseModel):
    inspection_id: Optional[int] = None
    equipment_id: int
    violation_type: str  # РўРёРї РЅР°СЂСѓС€РµРЅРёСЏ (РєСЂР°С‚РєРѕРµ РѕРїРёСЃР°РЅРёРµ РѕС‚ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ)
    context: Optional[str] = None


class AuditEventResponse(BaseModel):
    id: str
    entity_type: str
    entity_id: str
    action: str
    field_changes: Optional[Dict[str, Any]] = None
    performed_by: Optional[int] = None
    performed_at: datetime
    source: str
    trace_id: Optional[str] = None

class ViolationBulkCreate(BaseModel):
    equipment_ids: List[int]
    description: str
    inspection_id: Optional[int] = None
    fnp_clause: Optional[str] = None
    gost_clause: Optional[str] = None
    severity: str = "medium"
    location: Optional[str] = None
    deadline: Optional[datetime] = None

class ViolationBulkResponse(BaseModel):
    created: int
    skipped: int
    created_ids: List[int]
    errors: List[dict]

class ViolationBulkStatusUpdateRequest(BaseModel):
    violation_ids: List[int]
    status: str  # open, resolved
    resolved_by: Optional[int] = None

class ViolationBulkStatusUpdateResponse(BaseModel):
    updated: int
    errors: List[dict]

class ViolationSLARuleCreate(BaseModel):
    name: str
    violation_type: Optional[str] = None
    severity: Optional[str] = None
    days: int
    priority: int = 100
    is_active: bool = True

class ViolationSLARuleUpdate(BaseModel):
    name: Optional[str] = None
    violation_type: Optional[str] = None
    severity: Optional[str] = None
    days: Optional[int] = None
    priority: Optional[int] = None
    is_active: Optional[bool] = None

class ViolationSLARuleResponse(BaseModel):
    id: int
    name: str
    violation_type: Optional[str]
    severity: Optional[str]
    days: int
    priority: int
    is_active: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class ViolationSLAApplyRequest(BaseModel):
    limit: int = 200
    only_without_deadline: bool = True

@router.get("", response_model=List[ViolationResponse])
async def get_violations(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    equipment_id: Optional[int] = None,
    inspection_id: Optional[int] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    overdue: Optional[bool] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РџРѕР»СѓС‡РёС‚СЊ СЃРїРёСЃРѕРє РЅР°СЂСѓС€РµРЅРёР№"""
    await require_permission(current_user, "violations:read", db)
    
    query = select(Violation).options(selectinload(Violation.equipment))
    
    if equipment_id:
        query = query.where(Violation.equipment_id == equipment_id)
    
    if inspection_id:
        query = query.where(Violation.inspection_id == inspection_id)
    
    if status:
        query = query.where(Violation.status == status)
    
    if severity:
        query = query.where(Violation.severity == severity)
    if overdue is not None:
        now = datetime.utcnow()
        if overdue:
            query = query.where(
                Violation.status != "resolved",
                Violation.deadline.isnot(None),
                Violation.deadline < now
            )
        else:
            query = query.where(
                or_(
                    Violation.deadline.is_(None),
                    Violation.deadline >= now,
                    Violation.status == "resolved"
                )
            )
    
    query = query.order_by(Violation.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    violations = result.scalars().all()
    await _update_overdue_flags(db, violations)
    
    return [_violation_to_response(v) for v in violations]


@router.get("/export")
async def export_violations(
    equipment_id: Optional[int] = None,
    inspection_id: Optional[int] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    export_format: str = Query("xlsx", alias="format", regex="^(csv|xlsx)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Экспорт нарушений в CSV или XLSX."""
    await require_permission(current_user, "violations:read", db)

    query = select(Violation).options(selectinload(Violation.equipment))

    if equipment_id:
        query = query.where(Violation.equipment_id == equipment_id)
    if inspection_id:
        query = query.where(Violation.inspection_id == inspection_id)
    if status:
        query = query.where(Violation.status == status)
    if severity:
        query = query.where(Violation.severity == severity)

    result = await db.execute(query.order_by(Violation.created_at.desc()))
    violations = result.scalars().all()

    status_map = {
        "resolved": "Устранено",
        "open": "В работе",
        "in_progress": "В работе",
    }
    severity_map = {
        "critical": "Критическое",
        "high": "Высокое",
        "medium": "Среднее",
        "low": "Низкое",
    }

    def _format_date_ru(value: Optional[datetime]) -> str:
        if not value:
            return ""
        return value.strftime("%d.%m.%Y")

    columns = [
        ("id", "ID"),
        ("created_at", "Дата создания"),
        ("status", "Статус"),
        ("severity", "Критичность"),
        ("equipment_passport", "Паспорт оборудования"),
        ("equipment_position", "Позиция"),
        ("workshop", "Цех"),
        ("description", "Описание нарушения"),
        ("fnp_clause", "Пункт ФНП"),
        ("gost_clause", "Пункт ГОСТ"),
        ("deadline", "Срок устранения"),
        ("is_overdue", "Просрочено"),
    ]

    rows = []
    for violation in violations:
        equipment = getattr(violation, "equipment", None)
        rows.append({
            "id": violation.id,
            "created_at": _format_date_ru(violation.created_at),
            "status": status_map.get(violation.status, violation.status or ""),
            "severity": severity_map.get(violation.severity, violation.severity or ""),
            "equipment_passport": equipment.passport_number if equipment else "",
            "equipment_position": equipment.position if equipment else "",
            "workshop": equipment.workshop if equipment else "",
            "description": violation.description.replace("\n", " ") if violation.description else "",
            "fnp_clause": violation.fnp_clause or "",
            "gost_clause": violation.gost_clause or "",
            "deadline": _format_date_ru(violation.deadline),
            "is_overdue": "Да" if violation.is_overdue else "Нет",
        })

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if export_format == "csv":
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow([title for _, title in columns])
        for row in rows:
            writer.writerow([row[key] for key, _ in columns])
        filename = f"violations_{timestamp}.csv"
        return Response(
            content=output.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Для XLSX-экспорта требуется пакет openpyxl",
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Нарушения"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    status_fills = {
        "Устранено": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "В работе": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }
    severity_fills = {
        "Критическое": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Высокое": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
        "Среднее": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Низкое": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    }

    for col_idx, (_, title) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    widths = [8, 14, 14, 14, 22, 16, 14, 46, 16, 16, 16, 12]
    for idx, width in enumerate(widths, start=1):
        letter = chr(64 + idx)
        sheet.column_dimensions[letter].width = width

    status_col_idx = [i for i, (key, _) in enumerate(columns, start=1) if key == "status"][0]
    severity_col_idx = [i for i, (key, _) in enumerate(columns, start=1) if key == "severity"][0]
    overdue_col_idx = [i for i, (key, _) in enumerate(columns, start=1) if key == "is_overdue"][0]

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = border
            cell.alignment = body_alignment
            if row_idx % 2 == 0:
                cell.fill = even_fill

        status_cell = sheet.cell(row=row_idx, column=status_col_idx)
        if row_data["status"] in status_fills:
            status_cell.fill = status_fills[row_data["status"]]

        severity_cell = sheet.cell(row=row_idx, column=severity_col_idx)
        if row_data["severity"] in severity_fills:
            severity_cell.fill = severity_fills[row_data["severity"]]

        overdue_cell = sheet.cell(row=row_idx, column=overdue_col_idx)
        overdue_cell.fill = PatternFill(
            start_color="FEE2E2" if row_data["is_overdue"] == "Да" else "DCFCE7",
            end_color="FEE2E2" if row_data["is_overdue"] == "Да" else "DCFCE7",
            fill_type="solid",
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{max(2, len(rows) + 1)}"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"violations_{timestamp}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
@router.get("/{violation_id}", response_model=ViolationResponse)
async def get_violation(
    violation_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РџРѕР»СѓС‡РёС‚СЊ РЅР°СЂСѓС€РµРЅРёРµ РїРѕ ID"""
    await require_permission(current_user, "violations:read", db)
    
    result = await db.execute(
        select(Violation)
        .options(selectinload(Violation.equipment))
        .where(Violation.id == violation_id)
    )
    violation = result.scalar_one_or_none()
    
    if not violation:
        raise HTTPException(status_code=404, detail="Violation not found")
    await _update_overdue_flags(db, [violation])
    return _violation_to_response(violation)


@router.get("/{violation_id}/audit", response_model=List[AuditEventResponse])
async def get_violation_audit_history(
    violation_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await require_permission(current_user, "violations:read", db)

    result = await db.execute(
        select(AuditLog)
        .where(AuditLog.entity_type == "violation", AuditLog.entity_id == str(violation_id))
        .order_by(AuditLog.performed_at.desc())
    )
    rows = result.scalars().all()
    return [
        AuditEventResponse(
            id=row.id,
            entity_type=row.entity_type,
            entity_id=row.entity_id,
            action=row.action,
            field_changes=row.field_changes,
            performed_by=row.performed_by,
            performed_at=row.performed_at,
            source=row.source,
            trace_id=row.trace_id,
        )
        for row in rows
    ]

@router.post("", response_model=ViolationResponse, status_code=status.HTTP_201_CREATED)
async def create_violation(
    violation_data: ViolationCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РЎРѕР·РґР°С‚СЊ РЅРѕРІРѕРµ РЅР°СЂСѓС€РµРЅРёРµ"""
    await require_permission(current_user, "violations:create", db)
    
    # РџСЂРѕРІРµСЂРєР° СЃСѓС‰РµСЃС‚РІРѕРІР°РЅРёСЏ РѕР±РѕСЂСѓРґРѕРІР°РЅРёСЏ
    eq_result = await db.execute(select(Equipment).where(Equipment.id == violation_data.equipment_id))
    equipment = eq_result.scalar_one_or_none()
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    
    new_violation = Violation(
        inspection_id=violation_data.inspection_id,
        equipment_id=violation_data.equipment_id,
        description=violation_data.description,
        fnp_clause=violation_data.fnp_clause,
        gost_clause=violation_data.gost_clause,
        severity=violation_data.severity,
        criticality_level=violation_data.criticality_level,
        violation_type=violation_data.violation_type,
        violation_type_description=violation_data.violation_type_description,
        norm_reference=violation_data.norm_reference,
        recommended_act_text=violation_data.recommended_act_text,
        requirements=violation_data.requirements,
        source=violation_data.source,
        reported_by=violation_data.reported_by or current_user.id,
        attachment_meta=violation_data.attachment_meta,
        ai_classification=violation_data.ai_classification,
        ai_recommendations=violation_data.ai_recommendations,
        ai_payload_raw=violation_data.ai_payload_raw,
        location=violation_data.location,
        deadline=violation_data.deadline,
        status="open",
        created_by=current_user.id
    )
    if violation_data.deadline:
        new_violation.deadline_source = "manual"
        new_violation.deadline_rule_id = None
    db.add(new_violation)
    await db.flush()
    if not new_violation.deadline:
        await _apply_sla_deadline(db, new_violation)
    
    # Р›РѕРіРёСЂРѕРІР°РЅРёРµ
    activity = UserActivity(
        user_id=current_user.id,
        action_type="create",
        entity_type="violation",
        entity_id=new_violation.id,
        description=f"Created violation for equipment {violation_data.equipment_id}"
    )
    db.add(activity)

    await log_audit_event(
        db,
        entity_type="violation",
        entity_id=new_violation.id,
        action="CREATE",
        field_changes={
            "status": {"old": None, "new": new_violation.status},
            "severity": {"old": None, "new": new_violation.severity},
            "deadline": {"old": None, "new": new_violation.deadline.isoformat() if new_violation.deadline else None},
            "description": {"old": None, "new": new_violation.description},
        },
        performed_by=current_user.id,
        source=(new_violation.source or "ui"),
        trace_id=getattr(request.state, "trace_id", None),
    )

    await db.commit()
    await db.refresh(new_violation)
    new_violation.equipment = equipment
    
    return _violation_to_response(new_violation)


@router.post("/bulk", response_model=ViolationBulkResponse)
async def bulk_create_violations(
    payload: ViolationBulkCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РњР°СЃСЃРѕРІРѕРµ СЃРѕР·РґР°РЅРёРµ РЅР°СЂСѓС€РµРЅРёР№ РґР»СЏ РЅРµСЃРєРѕР»СЊРєРёС… РџРЎ"""
    await require_permission(current_user, "violations:create", db)

    if not payload.equipment_ids:
        raise HTTPException(status_code=400, detail="Equipment IDs are required")

    created_ids: List[int] = []
    errors: List[dict] = []

    for eq_id in payload.equipment_ids:
        eq_result = await db.execute(select(Equipment).where(Equipment.id == eq_id))
        equipment = eq_result.scalar_one_or_none()
        if not equipment:
            errors.append(
                {
                    "equipment_id": eq_id,
                    "detail": "Equipment not found",
                }
            )
            continue

        try:
            new_violation = Violation(
                inspection_id=payload.inspection_id,
                equipment_id=eq_id,
                description=payload.description,
                fnp_clause=payload.fnp_clause,
                gost_clause=payload.gost_clause,
                severity=payload.severity,
                location=payload.location,
                deadline=payload.deadline,
                status="open",
                created_by=current_user.id
            )
            if payload.deadline:
                new_violation.deadline_source = "manual"
                new_violation.deadline_rule_id = None
            db.add(new_violation)
            await db.flush()
            if not new_violation.deadline:
                await _apply_sla_deadline(db, new_violation)
            created_ids.append(new_violation.id)

            activity = UserActivity(
                user_id=current_user.id,
                action_type="create",
                entity_type="violation",
                entity_id=new_violation.id,
                description=f"Bulk created violation for equipment {eq_id}"
            )
            db.add(activity)
        except Exception as exc:
            errors.append(
                {
                    "equipment_id": eq_id,
                    "detail": str(exc),
                }
            )

    await db.commit()

    return ViolationBulkResponse(
        created=len(created_ids),
        skipped=len(payload.equipment_ids) - len(created_ids),
        created_ids=created_ids,
        errors=errors,
    )

@router.put("/bulk/status", response_model=ViolationBulkStatusUpdateResponse)
async def bulk_update_violation_status(
    payload: ViolationBulkStatusUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РњР°СЃСЃРѕРІРѕРµ РёР·РјРµРЅРµРЅРёРµ СЃС‚Р°С‚СѓСЃРѕРІ РЅР°СЂСѓС€РµРЅРёР№"""
    await require_permission(current_user, "violations:update", db)

    if not payload.violation_ids:
        raise HTTPException(status_code=400, detail="Violation IDs are required")

    if payload.status not in ["open", "resolved"]:
        raise HTTPException(status_code=400, detail="Status must be 'open' or 'resolved'")

    updated = 0
    errors: List[dict] = []

    for violation_id in payload.violation_ids:
        try:
            result = await db.execute(select(Violation).where(Violation.id == violation_id))
            violation = result.scalar_one_or_none()
            
            if not violation:
                errors.append({"violation_id": violation_id, "detail": "Violation not found"})
                continue

            old_status = violation.status
            violation.status = payload.status
            violation.updated_at = datetime.utcnow()

            if payload.status == "resolved" and not violation.resolved_at:
                violation.resolved_at = datetime.utcnow()
                violation.resolved_by = payload.resolved_by or current_user.id
                violation.is_overdue = False
                violation.overdue_at = None
            elif payload.status == "open":
                violation.resolved_at = None
                violation.resolved_by = None

            # Р›РѕРіРёСЂРѕРІР°РЅРёРµ
            activity = UserActivity(
                user_id=current_user.id,
                action_type="update",
                entity_type="violation",
                entity_id=violation.id,
                description=f"Bulk updated violation status from {old_status} to {payload.status}"
            )
            db.add(activity)
            updated += 1

        except Exception as exc:
            errors.append({"violation_id": violation_id, "detail": str(exc)})

    await db.commit()

    return ViolationBulkStatusUpdateResponse(updated=updated, errors=errors)

@router.get("/sla-rules", response_model=List[ViolationSLARuleResponse])
async def get_sla_rules(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    await require_permission(current_user, "violations:read", db)
    result = await db.execute(select(ViolationSLARule).order_by(ViolationSLARule.priority.asc()))
    rules = result.scalars().all()
    return [
        ViolationSLARuleResponse(
            id=r.id,
            name=r.name,
            violation_type=r.violation_type,
            severity=r.severity,
            days=r.days,
            priority=r.priority,
            is_active=r.is_active,
            created_at=r.created_at,
            updated_at=r.updated_at,
        )
        for r in rules
    ]

@router.post("/sla-rules", response_model=ViolationSLARuleResponse, status_code=status.HTTP_201_CREATED)
async def create_sla_rule(
    payload: ViolationSLARuleCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    await require_permission(current_user, "violations:update", db)
    rule = ViolationSLARule(
        name=payload.name,
        violation_type=payload.violation_type,
        severity=payload.severity,
        days=payload.days,
        priority=payload.priority,
        is_active=payload.is_active
    )
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return ViolationSLARuleResponse(
        id=rule.id,
        name=rule.name,
        violation_type=rule.violation_type,
        severity=rule.severity,
        days=rule.days,
        priority=rule.priority,
        is_active=rule.is_active,
        created_at=rule.created_at,
        updated_at=rule.updated_at,
    )

@router.put("/sla-rules/{rule_id}", response_model=ViolationSLARuleResponse)
async def update_sla_rule(
    rule_id: int,
    payload: ViolationSLARuleUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    await require_permission(current_user, "violations:update", db)
    result = await db.execute(select(ViolationSLARule).where(ViolationSLARule.id == rule_id))
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(status_code=404, detail="SLA rule not found")
    update_data = payload.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(rule, field, value)
    rule.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(rule)
    return ViolationSLARuleResponse(
        id=rule.id,
        name=rule.name,
        violation_type=rule.violation_type,
        severity=rule.severity,
        days=rule.days,
        priority=rule.priority,
        is_active=rule.is_active,
        created_at=rule.created_at,
        updated_at=rule.updated_at,
    )

@router.delete("/sla-rules/{rule_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_sla_rule(
    rule_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    await require_permission(current_user, "violations:update", db)
    result = await db.execute(select(ViolationSLARule).where(ViolationSLARule.id == rule_id))
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(status_code=404, detail="SLA rule not found")
    await db.delete(rule)
    await db.commit()
    return None

@router.post("/sla/apply")
async def apply_sla_to_violations(
    payload: ViolationSLAApplyRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    await require_permission(current_user, "violations:update", db)
    limit = max(1, min(payload.limit, 2000))
    query = select(Violation).where(Violation.status != "resolved")
    if payload.only_without_deadline:
        query = query.where(Violation.deadline.is_(None))
    result = await db.execute(query.limit(limit))
    violations = result.scalars().all()
    updated = 0
    for violation in violations:
        before = violation.deadline
        await _apply_sla_deadline(db, violation)
        if before != violation.deadline:
            updated += 1
    await db.commit()
    return {"status": "ok", "processed": len(violations), "updated": updated}

@router.post("/ai/generate", response_model=AIGenerateViolationResponse)
async def generate_violation_ai(
    request: AIGenerateViolationRequest,
    http_request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Р“РµРЅРµСЂР°С†РёСЏ РЅР°СЂСѓС€РµРЅРёСЏ С‡РµСЂРµР· РР"""
    await require_permission(current_user, "violations:create", db)
    
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        # РСЃРїРѕР»СЊР·СѓРµРј СѓРЅРёРІРµСЂСЃР°Р»СЊРЅС‹Р№ AI РєР»РёРµРЅС‚
        try:
            from backend.ai_client import get_ai_client_async, AITemporarilyUnavailableError
        except ImportError:
            from ai_client import get_ai_client_async, AITemporarilyUnavailableError
        
        # Р—Р°РіСЂСѓР¶Р°РµРј РЅР°СЃС‚СЂРѕР№РєРё РёР· Р‘Р”
        logger.info(f"Р—Р°РіСЂСѓР·РєР° AI РєР»РёРµРЅС‚Р° РґР»СЏ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ {current_user.id}")
        ai_client = await get_ai_client_async(db)
        if not ai_client:
            logger.error("AI РєР»РёРµРЅС‚ РЅРµ РЅР°СЃС‚СЂРѕРµРЅ")
            raise HTTPException(
                status_code=400, 
                detail="AI РЅРµ РЅР°СЃС‚СЂРѕРµРЅ. РџРµСЂРµР№РґРёС‚Рµ РІ СЂР°Р·РґРµР» 'РќР°СЃС‚СЂРѕР№РєРё' -> 'РЎРёСЃС‚РµРјРЅС‹Рµ РЅР°СЃС‚СЂРѕР№РєРё' Рё РЅР°СЃС‚СЂРѕР№С‚Рµ AI РїСЂРѕРІР°Р№РґРµСЂР°."
            )
        
        logger.info(f"AI РєР»РёРµРЅС‚ Р·Р°РіСЂСѓР¶РµРЅ, РїСЂРѕРІР°Р№РґРµСЂ: {ai_client.provider}")
        
        # РџРѕР»СѓС‡РµРЅРёРµ РёРЅС„РѕСЂРјР°С†РёРё РѕР± РѕР±РѕСЂСѓРґРѕРІР°РЅРёРё
        eq_result = await db.execute(select(Equipment).where(Equipment.id == request.equipment_id))
        equipment = eq_result.scalar_one_or_none()
        if not equipment:
            logger.error(f"РћР±РѕСЂСѓРґРѕРІР°РЅРёРµ СЃ ID {request.equipment_id} РЅРµ РЅР°Р№РґРµРЅРѕ")
            raise HTTPException(status_code=404, detail="Equipment not found")
        
        logger.info(f"Р“РµРЅРµСЂР°С†РёСЏ РЅР°СЂСѓС€РµРЅРёСЏ РґР»СЏ РѕР±РѕСЂСѓРґРѕРІР°РЅРёСЏ {equipment.id}")
        
        # РџРѕР»СѓС‡Р°РµРј СЂРµР»РµРІР°РЅС‚РЅСѓСЋ РёРЅС„РѕСЂРјР°С†РёСЋ РёР· Р±Р°Р·С‹ Р·РЅР°РЅРёР№
        knowledge_context = ""
        used_documents = []  # РЎРїРёСЃРѕРє РёСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹С… РґРѕРєСѓРјРµРЅС‚РѕРІ РґР»СЏ РѕС‚РІРµС‚Р°
        has_knowledge_sources = False
        try:
            try:
                from backend.models import KnowledgeBase
            except ImportError:
                from ..models import KnowledgeBase
            
            # РџР РРћР РРўР•Рў: РЎРЅР°С‡Р°Р»Р° РёС‰РµРј Р¤РќРџ 461/Р“РћРЎРў, Р·Р°С‚РµРј Р»СЋР±С‹Рµ РґРѕРєСѓРјРµРЅС‚С‹.
            # Р’Р°Р¶РЅРѕ: РґРѕРєСѓРјРµРЅС‚С‹ РјРѕРіСѓС‚ Р±С‹С‚СЊ Р·Р°РіСЂСѓР¶РµРЅС‹ РєР°Рє "other"/"manual", РїРѕСЌС‚РѕРјСѓ РґРµР»Р°РµРј РєР°СЃРєР°РґРЅС‹Р№ РїРѕРёСЃРє.
            search_terms = [
                (request.violation_type or "").strip().lower(),
                (equipment.equipment_type or "").strip().lower(),
                (request.context or "").strip().lower(),
            ]
            search_terms = [term for term in search_terms if term and len(term) >= 2]

            def build_search_conditions(include_clause_number: bool = True):
                conditions = []
                for term in search_terms:
                    term_filters = [
                        KnowledgeBase.title.ilike(f"%{term}%"),
                        KnowledgeBase.content.ilike(f"%{term}%"),
                        KnowledgeBase.section.ilike(f"%{term}%"),
                    ]
                    if include_clause_number:
                        term_filters.append(KnowledgeBase.clause_number.ilike(f"%{term}%"))
                    conditions.append(or_(*term_filters))
                return conditions

            fnp_gost_query = select(KnowledgeBase).where(
                KnowledgeBase.document_type.in_(["fnp461", "gost"])
            )
            fnp_conditions = build_search_conditions(include_clause_number=True)
            if fnp_conditions:
                fnp_gost_query = fnp_gost_query.where(or_(*fnp_conditions))
            fnp_gost_query = fnp_gost_query.limit(10)

            result = await db.execute(fnp_gost_query)
            knowledge_items = result.scalars().all()

            # 2) Р•СЃР»Рё РЅРµС‚ Р¤РќРџ/Р“РћРЎРў - Р±РµСЂРµРј Р»СЋР±РѕР№ С‚РёРї РґРѕРєСѓРјРµРЅС‚Р° СЃ С‚РµРјРё Р¶Рµ С‚РµСЂРјРёРЅР°РјРё.
            if not knowledge_items:
                logger.info("Р РµР»РµРІР°РЅС‚РЅС‹Рµ Р¤РќРџ/Р“РћРЎРў РЅРµ РЅР°Р№РґРµРЅС‹, РёС‰РµРј РІСЃРµ С‚РёРїС‹ РґРѕРєСѓРјРµРЅС‚РѕРІ")
                general_query = select(KnowledgeBase)
                general_conditions = build_search_conditions(include_clause_number=False)
                if general_conditions:
                    general_query = general_query.where(or_(*general_conditions))
                general_query = general_query.limit(10)
                result = await db.execute(general_query)
                knowledge_items = result.scalars().all()

            # 3) РџРѕСЃР»РµРґРЅРёР№ fallback: РµСЃР»Рё РїРѕ РєР»СЋС‡РµРІС‹Рј СЃР»РѕРІР°Рј РЅРёС‡РµРіРѕ РЅРµ РЅР°С€Р»Рё,
            # РІСЃРµ СЂР°РІРЅРѕ РїРµСЂРµРґР°РµРј РІ РР С‡Р°СЃС‚СЊ Р±Р°Р·С‹ Р·РЅР°РЅРёР№, С‡С‚РѕР±С‹ РЅРµ Р±С‹Р»Рѕ Р»РѕР¶РЅРѕРіРѕ "Р±Р°Р·Р° РїСѓСЃС‚Р°".
            if not knowledge_items:
                logger.info("РџРѕ РєР»СЋС‡РµРІС‹Рј СЃР»РѕРІР°Рј РЅРµС‚ СЃРѕРІРїР°РґРµРЅРёР№, Р±РµСЂРµРј РїРѕСЃР»РµРґРЅРёРµ РґРѕРєСѓРјРµРЅС‚С‹ РёР· Р±Р°Р·С‹")
                fallback_query = select(KnowledgeBase).order_by(KnowledgeBase.updated_at.desc()).limit(10)
                result = await db.execute(fallback_query)
                knowledge_items = result.scalars().all()
            
            if knowledge_items:
                has_knowledge_sources = True
                knowledge_context = "\n\n=== Р Р•Р›Р•Р’РђРќРўРќРђРЇ Р”РћРљРЈРњР•РќРўРђР¦РРЇ РР— Р‘РђР—Р« Р—РќРђРќРР™ ===\n"
                logger.info(f"РќР°Р№РґРµРЅРѕ {len(knowledge_items)} РґРѕРєСѓРјРµРЅС‚РѕРІ РІ Р±Р°Р·Рµ Р·РЅР°РЅРёР№")
                
                # Р¤СѓРЅРєС†РёСЏ РґР»СЏ СѓРјРЅРѕРіРѕ РёР·РІР»РµС‡РµРЅРёСЏ СЂРµР»РµРІР°РЅС‚РЅС‹С… С‡Р°СЃС‚РµР№ РґРѕРєСѓРјРµРЅС‚Р°
                def extract_relevant_content(content: str, search_terms: list, max_length: int = 10000) -> str:
                    """РР·РІР»РµРєР°РµС‚ СЂРµР»РµРІР°РЅС‚РЅС‹Рµ С‡Р°СЃС‚Рё РґРѕРєСѓРјРµРЅС‚Р° РІРѕРєСЂСѓРі РєР»СЋС‡РµРІС‹С… СЃР»РѕРІ"""
                    if not search_terms or not any(term.strip() for term in search_terms):
                        # Р•СЃР»Рё РЅРµС‚ РїРѕРёСЃРєРѕРІС‹С… С‚РµСЂРјРёРЅРѕРІ, Р±РµСЂРµРј РЅР°С‡Р°Р»Рѕ РґРѕРєСѓРјРµРЅС‚Р°
                        return content[:max_length] + ("..." if len(content) > max_length else "")
                    
                    content_lower = content.lower()
                    relevant_parts = []
                    used_positions = set()
                    
                    # РС‰РµРј РІС…РѕР¶РґРµРЅРёСЏ РєР°Р¶РґРѕРіРѕ С‚РµСЂРјРёРЅР°
                    for term in search_terms:
                        if not term or not term.strip():
                            continue
                        term_lower = term.lower().strip()
                        if len(term_lower) < 3:  # РџСЂРѕРїСѓСЃРєР°РµРј СЃР»РёС€РєРѕРј РєРѕСЂРѕС‚РєРёРµ С‚РµСЂРјРёРЅС‹
                            continue
                        
                        # РС‰РµРј РІСЃРµ РІС…РѕР¶РґРµРЅРёСЏ С‚РµСЂРјРёРЅР°
                        start = 0
                        while True:
                            pos = content_lower.find(term_lower, start)
                            if pos == -1:
                                break
                            
                            # РР·РІР»РµРєР°РµРј РєРѕРЅС‚РµРєСЃС‚ РІРѕРєСЂСѓРі РЅР°Р№РґРµРЅРЅРѕРіРѕ С‚РµСЂРјРёРЅР° (2000 СЃРёРјРІРѕР»РѕРІ РґРѕ Рё РїРѕСЃР»Рµ)
                            context_start = max(0, pos - 2000)
                            context_end = min(len(content), pos + len(term_lower) + 2000)
                            
                            # РџСЂРѕРІРµСЂСЏРµРј, РЅРµ РїРµСЂРµСЃРµРєР°РµС‚СЃСЏ Р»Рё СЃ СѓР¶Рµ РёСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹РјРё С‡Р°СЃС‚СЏРјРё
                            overlap = False
                            for used_start, used_end in used_positions:
                                if not (context_end < used_start or context_start > used_end):
                                    overlap = True
                                    break
                            
                            if not overlap:
                                relevant_parts.append((context_start, context_end))
                                used_positions.add((context_start, context_end))
                            
                            start = pos + 1
                            
                            # РћРіСЂР°РЅРёС‡РёРІР°РµРј РєРѕР»РёС‡РµСЃС‚РІРѕ РЅР°Р№РґРµРЅРЅС‹С… РІС…РѕР¶РґРµРЅРёР№
                            if len(relevant_parts) >= 5:
                                break
                    
                    if relevant_parts:
                        # РЎРѕСЂС‚РёСЂСѓРµРј РїРѕ РїРѕР·РёС†РёРё Рё РѕР±СЉРµРґРёРЅСЏРµРј
                        relevant_parts.sort()
                        result_parts = []
                        current_start, current_end = relevant_parts[0]
                        
                        for start, end in relevant_parts[1:]:
                            if start <= current_end + 500:  # РћР±СЉРµРґРёРЅСЏРµРј Р±Р»РёР·РєРёРµ С‡Р°СЃС‚Рё
                                current_end = max(current_end, end)
                            else:
                                result_parts.append((current_start, current_end))
                                current_start, current_end = start, end
                        result_parts.append((current_start, current_end))
                        
                        # РћР±СЉРµРґРёРЅСЏРµРј С‡Р°СЃС‚Рё
                        extracted = []
                        for start, end in result_parts:
                            part = content[start:end]
                            if start > 0:
                                extracted.append("...")
                            extracted.append(part)
                            if end < len(content):
                                extracted.append("...")
                        
                        result = "".join(extracted)
                        # РћРіСЂР°РЅРёС‡РёРІР°РµРј РѕР±С‰СѓСЋ РґР»РёРЅСѓ
                        if len(result) > max_length:
                            result = result[:max_length] + "\n[Р”РѕРєСѓРјРµРЅС‚ РїСЂРѕРґРѕР»Р¶Р°РµС‚СЃСЏ, РїРѕРєР°Р·Р°РЅС‹ СЂРµР»РµРІР°РЅС‚РЅС‹Рµ С‡Р°СЃС‚Рё]"
                        return result
                    else:
                        # Р•СЃР»Рё РЅРµ РЅР°С€Р»Рё СЂРµР»РµРІР°РЅС‚РЅС‹С… С‡Р°СЃС‚РµР№, Р±РµСЂРµРј РЅР°С‡Р°Р»Рѕ РґРѕРєСѓРјРµРЅС‚Р°
                        return content[:max_length] + ("..." if len(content) > max_length else "")
                
                # РћР±СЂР°Р±Р°С‚С‹РІР°РµРј РґРѕРєСѓРјРµРЅС‚С‹
                for item in knowledge_items[:6]:  # Ограничиваем количество документов для стабильной генерации
                    doc_type_name = {
                        "fnp461": "Р¤РќРџ 461",
                        "gost": "Р“РћРЎРў",
                        "manual": "РњРµС‚РѕРґРёС‡РєР°"
                    }.get(item.document_type, item.document_type.upper())
                    
                    knowledge_context += f"\n{'='*60}\n[{doc_type_name}] {item.title}\n{'='*60}\n"
                    if item.section:
                        knowledge_context += f"Р Р°Р·РґРµР»: {item.section}\n"
                    if item.clause_number:
                        knowledge_context += f"РџСѓРЅРєС‚: {item.clause_number}\n"
                    knowledge_context += "\n"
                    
                    # РЈРјРЅРѕРµ РёР·РІР»РµС‡РµРЅРёРµ РєРѕРЅС‚РµРЅС‚Р°:
                    # - Р¤РќРџ/Р“РћРЎРў: РґРѕ 10000 СЃРёРјРІРѕР»РѕРІ СЂРµР»РµРІР°РЅС‚РЅС‹С… С‡Р°СЃС‚РµР№ РёР»Рё РІРµСЃСЊ РґРѕРєСѓРјРµРЅС‚
                    # - РњРµС‚РѕРґРёС‡РєРё: РґРѕ 5000 СЃРёРјРІРѕР»РѕРІ
                    if item.document_type in ["fnp461", "gost"]:
                        # Р”Р»СЏ Р¤РќРџ/Р“РћРЎРў РёСЃРїРѕР»СЊР·СѓРµРј СѓРјРЅРѕРµ РёР·РІР»РµС‡РµРЅРёРµ СЂРµР»РµРІР°РЅС‚РЅС‹С… С‡Р°СЃС‚РµР№
                        content_preview = extract_relevant_content(
                            item.content, 
                            search_terms, 
                            max_length=3000  # Ограничиваем размер контекста на документ
                        )
                    else:
                        # Р”Р»СЏ РјРµС‚РѕРґРёС‡РµРє С‚РѕР¶Рµ РёСЃРїРѕР»СЊР·СѓРµРј СѓРјРЅРѕРµ РёР·РІР»РµС‡РµРЅРёРµ, РЅРѕ РјРµРЅСЊС€Рµ
                        content_preview = extract_relevant_content(
                            item.content,
                            search_terms,
                            max_length=1800
                        )
                    
                    knowledge_context += f"{content_preview}\n\n"
                    
                    # РЎРѕС…СЂР°РЅСЏРµРј РёРЅС„РѕСЂРјР°С†РёСЋ Рѕ РґРѕРєСѓРјРµРЅС‚Рµ РґР»СЏ РѕС‚РІРµС‚Р°
                    used_documents.append({
                        "id": item.id,
                        "document_type": item.document_type,
                        "title": item.title,
                        "section": item.section,
                        "clause_number": item.clause_number,
                        "content_preview": content_preview[:200] + "..." if len(content_preview) > 200 else content_preview
                    })
                    
                    logger.info(f"Р”РѕР±Р°РІР»РµРЅ РґРѕРєСѓРјРµРЅС‚ РІ РєРѕРЅС‚РµРєСЃС‚: {doc_type_name} - {item.title} (ID: {item.id})")

                max_context_chars = 18000
                if len(knowledge_context) > max_context_chars:
                    logger.info(f"Ограничиваем knowledge_context: {len(knowledge_context)} -> {max_context_chars} символов")
                    knowledge_context = knowledge_context[:max_context_chars] + "\n\n[Контекст обрезан для стабильности запроса к AI]"
            else:
                logger.warning("Р‘Р°Р·Р° Р·РЅР°РЅРёР№ РїСѓСЃС‚Р° РёР»Рё РЅРµ СЃРѕРґРµСЂР¶РёС‚ СЂРµР»РµРІР°РЅС‚РЅС‹С… РґРѕРєСѓРјРµРЅС‚РѕРІ. РР Р±СѓРґРµС‚ РіРµРЅРµСЂРёСЂРѕРІР°С‚СЊ РїСѓРЅРєС‚С‹ Р¤РќРџ Р±РµР· РєРѕРЅС‚РµРєСЃС‚Р°.")
                knowledge_context = "\n\nвљ пёЏ Р’РќРРњРђРќРР•: Р‘Р°Р·Р° Р·РЅР°РЅРёР№ РЅРµ СЃРѕРґРµСЂР¶РёС‚ РґРѕРєСѓРјРµРЅС‚РѕРІ Р¤РќРџ 461 РёР»Рё Р“РћРЎРў. РСЃРїРѕР»СЊР·СѓР№ С‚РѕР»СЊРєРѕ Р Р•РђР›Р¬РќР«Р• РїСѓРЅРєС‚С‹, РєРѕС‚РѕСЂС‹Рµ С‚С‹ Р·РЅР°РµС€СЊ. Р•СЃР»Рё РЅРµ СѓРІРµСЂРµРЅ - СѓРєР°Р¶Рё 'РЅРµ РїСЂРёРјРµРЅРёРјРѕ'.\n"
        except Exception as e:
            logger.warning(f"РќРµ СѓРґР°Р»РѕСЃСЊ Р·Р°РіСЂСѓР·РёС‚СЊ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№: {e}")
            knowledge_context = "\n\nвљ пёЏ РћС€РёР±РєР° Р·Р°РіСЂСѓР·РєРё Р±Р°Р·С‹ Р·РЅР°РЅРёР№. РСЃРїРѕР»СЊР·СѓР№ С‚РѕР»СЊРєРѕ Р Р•РђР›Р¬РќР«Р• РїСѓРЅРєС‚С‹ Р¤РќРџ 461/Р“РћРЎРў.\n"
        
        prompt = f"""РћС„РѕСЂРјРё РѕС„РёС†РёР°Р»СЊРЅРѕРµ РЅР°СЂСѓС€РµРЅРёРµ РґР»СЏ РїРѕРґСЉРµРјРЅРѕРіРѕ СЃРѕРѕСЂСѓР¶РµРЅРёСЏ РЅР° РѕСЃРЅРѕРІРµ С‚РёРїР° РЅР°СЂСѓС€РµРЅРёСЏ.

РўРРџ РќРђР РЈРЁР•РќРРЇ (РѕС‚ РёРЅСЃРїРµРєС‚РѕСЂР°): {request.violation_type}

Р”РђРќРќР«Р• РћР‘РћР РЈР”РћР’РђРќРРЇ:
- РўРёРї РџРЎ: {equipment.equipment_type}
- РџР°СЃРїРѕСЂС‚: {equipment.passport_number}
- РњРµСЃС‚Рѕ СѓСЃС‚Р°РЅРѕРІРєРё: {equipment.installation_location or 'РќРµ СѓРєР°Р·Р°РЅРѕ'}
- РљРѕРЅС‚РµРєСЃС‚: {request.context or 'РќРµ СѓРєР°Р·Р°РЅРѕ'}

{knowledge_context}

Р—РђР”РђР§Рђ:
1. РЎРѕР·РґР°Р№ РћР¤РР¦РРђР›Р¬РќРћР•, Р”РћРљРЈРњР•РќРўРђР›Р¬РќРћР• РѕРїРёСЃР°РЅРёРµ РЅР°СЂСѓС€РµРЅРёСЏ РІ РѕС„РёС†РёР°Р»СЊРЅРѕРј СЃС‚РёР»Рµ РёРЅСЃРїРµРєС†РёРё (2-4 РїСЂРµРґР»РѕР¶РµРЅРёСЏ)
2. РћРїСЂРµРґРµР»Рё Рё СѓРєР°Р¶Рё РєРѕРЅРєСЂРµС‚РЅС‹Р№ РїСѓРЅРєС‚ Р¤РќРџ 461, РєРѕС‚РѕСЂС‹Р№ РЅР°СЂСѓС€РµРЅ (С„РѕСЂРјР°С‚: "Рї. 123 Р¤РќРџ 461" РёР»Рё "Рї.Рї. 123-125 Р¤РќРџ 461")
3. РћРїСЂРµРґРµР»Рё Рё СѓРєР°Р¶Рё Р“РћРЎРў, РµСЃР»Рё РїСЂРёРјРµРЅРёРјРѕ (С„РѕСЂРјР°С‚: "Р“РћРЎРў 12345-2020" РёР»Рё "Р“РћРЎРў 12345")
4. РћРїСЂРµРґРµР»Рё СЃСЂРѕРє СѓСЃС‚СЂР°РЅРµРЅРёСЏ РЅР° РѕСЃРЅРѕРІРµ РєСЂРёС‚РёС‡РЅРѕСЃС‚Рё (С„РѕСЂРјР°С‚: РєРѕР»РёС‡РµСЃС‚РІРѕ РґРЅРµР№, РЅР°РїСЂРёРјРµСЂ "30" РґР»СЏ СЃСЂРµРґРЅРёС… РЅР°СЂСѓС€РµРЅРёР№)

Р¤РћР РњРђРў РћРўР’Р•РўРђ (СЃС‚СЂРѕРіРѕ СЃРѕР±Р»СЋРґР°Р№ СЃС‚СЂСѓРєС‚СѓСЂСѓ):
РћРџРРЎРђРќРР•: [РѕС„РёС†РёР°Р»СЊРЅРѕРµ РѕРїРёСЃР°РЅРёРµ РЅР°СЂСѓС€РµРЅРёСЏ 2-4 РїСЂРµРґР»РѕР¶РµРЅРёСЏ]
Р¤РќРџ: [РїСѓРЅРєС‚ Р¤РќРџ 461, РЅР°РїСЂРёРјРµСЂ "Рї. 123 Р¤РќРџ 461" РёР»Рё "РЅРµ РїСЂРёРјРµРЅРёРјРѕ"]
Р“РћРЎРў: [РЅРѕРјРµСЂ Р“РћРЎРў, РЅР°РїСЂРёРјРµСЂ "Р“РћРЎРў 12345-2020" РёР»Рё "РЅРµ РїСЂРёРјРµРЅРёРјРѕ"]
РЎР РћРљ_Р”РќР•Р™: [РєРѕР»РёС‡РµСЃС‚РІРѕ РґРЅРµР№ РґР»СЏ СѓСЃС‚СЂР°РЅРµРЅРёСЏ, РЅР°РїСЂРёРјРµСЂ "30"]

РўСЂРµР±РѕРІР°РЅРёСЏ:
- РћРїРёСЃР°РЅРёРµ РґРѕР»Р¶РЅРѕ Р±С‹С‚СЊ РћР¤РР¦РРђР›Р¬РќР«Рњ Рё Р”РћРљРЈРњР•РќРўРђР›Р¬РќР«Рњ
- РСЃРїРѕР»СЊР·СѓР№ РѕС„РёС†РёР°Р»СЊРЅСѓСЋ С‚РµСЂРјРёРЅРѕР»РѕРіРёСЋ РёРЅСЃРїРµРєС†РёРё
- РџСѓРЅРєС‚С‹ Р¤РќРџ/Р“РћРЎРў РґРѕР»Р¶РЅС‹ Р±С‹С‚СЊ Р Р•РђР›Р¬РќР«РњР Рё Р Р•Р›Р•Р’РђРќРўРќР«РњР
- РЎСЂРѕРє СѓСЃС‚СЂР°РЅРµРЅРёСЏ: РєСЂРёС‚РёС‡РЅС‹Рµ - 7 РґРЅРµР№, РІС‹СЃРѕРєРёРµ - 15 РґРЅРµР№, СЃСЂРµРґРЅРёРµ - 30 РґРЅРµР№, РЅРёР·РєРёРµ - 60 РґРЅРµР№"""
        
        system_prompt = "РўС‹ РїРѕРјРѕС‰РЅРёРє РґР»СЏ СЃРѕР·РґР°РЅРёСЏ РѕС„РёС†РёР°Р»СЊРЅС‹С… РЅР°СЂСѓС€РµРЅРёР№ РІ СЃРёСЃС‚РµРјРµ РёРЅСЃРїРµРєС†РёРё. РћС‚РІРµС‚ С‚РѕР»СЊРєРѕ РЅР° СЂСѓСЃСЃРєРѕРј. РўРІРѕСЏ Р·Р°РґР°С‡Р° - РѕС„РѕСЂРјРёС‚СЊ С‚РёРї РЅР°СЂСѓС€РµРЅРёСЏ РІ РѕС„РёС†РёР°Р»СЊРЅС‹Р№ РґРѕРєСѓРјРµРЅС‚ СЃ СѓРєР°Р·Р°РЅРёРµРј РїСѓРЅРєС‚РѕРІ Р¤РќРџ 461, Р“РћРЎРў Рё СЃСЂРѕРєР° СѓСЃС‚СЂР°РЅРµРЅРёСЏ. РћС‚РІРµС‡Р°Р№ СЃС‚СЂРѕРіРѕ РІ СѓРєР°Р·Р°РЅРЅРѕРј С„РѕСЂРјР°С‚Рµ."
        if not has_knowledge_sources:
            prompt += """

ОГРАНИЧЕНИЕ ПО ИСТОЧНИКАМ:
- В базе знаний не найдено релевантных нормативов.
- Не придумывай пункты ФНП/ГОСТ и номера документов.
- Для полей ФНП и ГОСТ верни строго: "не применимо".
- В описании явно укажи: "В базе знаний не найдено релевантных нормативов. Уточните запрос."
"""
        
        # Р›РѕРіРёСЂСѓРµРј РєРѕРЅС‚РµРєСЃС‚, РєРѕС‚РѕСЂС‹Р№ Р±С‹Р» РїРµСЂРµРґР°РЅ РР
        logger.info("=" * 80)
        logger.info("=== РљРћРќРўР•РљРЎРў Р”Р›РЇ РР ===")
        logger.info(f"РўРёРї РЅР°СЂСѓС€РµРЅРёСЏ: {request.violation_type}")
        logger.info(f"РћР±РѕСЂСѓРґРѕРІР°РЅРёРµ: {equipment.equipment_type} (ID: {equipment.id})")
        logger.info(f"РќР°Р№РґРµРЅРѕ РґРѕРєСѓРјРµРЅС‚РѕРІ РІ Р±Р°Р·Рµ Р·РЅР°РЅРёР№: {len(used_documents)}")
        if used_documents:
            logger.info("РСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹Рµ РґРѕРєСѓРјРµРЅС‚С‹:")
            for doc in used_documents:
                logger.info(f"  - {doc['document_type']}: {doc['title']} (РїСѓРЅРєС‚: {doc.get('clause_number', 'РЅ/Рґ')})")
        else:
            logger.warning("вљ пёЏ Р”РѕРєСѓРјРµРЅС‚С‹ РІ Р±Р°Р·Рµ Р·РЅР°РЅРёР№ РќР• РќРђР™Р”Р•РќР«!")
        logger.info(f"Р”Р»РёРЅР° РєРѕРЅС‚РµРєСЃС‚Р° Р±Р°Р·С‹ Р·РЅР°РЅРёР№: {len(knowledge_context)} СЃРёРјРІРѕР»РѕРІ")
        logger.info("=" * 80)
        
        logger.info("РћС‚РїСЂР°РІРєР° Р·Р°РїСЂРѕСЃР° Рє AI РґР»СЏ РіРµРЅРµСЂР°С†РёРё РЅР°СЂСѓС€РµРЅРёСЏ")

        ai_max_tokens = 1200
        try:
            max_tokens_result = await db.execute(
                select(SystemSettings.value).where(SystemSettings.key == "ai_max_tokens")
            )
            max_tokens_raw = max_tokens_result.scalar_one_or_none()
            if max_tokens_raw is not None:
                ai_max_tokens = max(200, min(int(str(max_tokens_raw).strip()), 4000))
        except Exception as settings_error:
            logger.warning(f"Не удалось прочитать ai_max_tokens из системных настроек: {settings_error}")

        retry_max_tokens = max(300, min(ai_max_tokens // 2, 1000))
        ai_generation_warning: Optional[str] = None
        try:
            # Для Timeweb Cloud не передаем temperature (некоторые модели не поддерживают)
            temperature = None if ai_client.provider == "timeweb" else 0.7

            ai_description = ai_client.generate_text(
                prompt=prompt,
                system_prompt=system_prompt,
                max_tokens=ai_max_tokens,
                temperature=temperature
            )
            logger.info(f"AI вернул ответ длиной {len(ai_description) if ai_description else 0} символов")
        except Exception as ai_error:
            logger.warning(f"Первичная генерация через AI не удалась: {str(ai_error)}")
            retry_error = None
            try:
                retry_prompt = f"""Оформи официальное нарушение в 3-4 предложениях.

ТИП НАРУШЕНИЯ: {request.violation_type}
ОБОРУДОВАНИЕ: {equipment.equipment_type}, паспорт {equipment.passport_number}
КОНТЕКСТ: {request.context or 'не указан'}

ФОРМАТ ОТВЕТА:
ОПИСАНИЕ: ...
ФНП: не применимо
ГОСТ: не применимо
СРОК_ДНЕЙ: 30"""
                ai_description = ai_client.generate_text(
                    prompt=retry_prompt,
                    system_prompt=system_prompt,
                    max_tokens=retry_max_tokens,
                    temperature=temperature
                )
                logger.info("AI генерация прошла со второй попытки (облегченный prompt)")
            except Exception as second_error:
                retry_error = second_error

            if retry_error is not None:
                logger.error(f"Ошибка при генерации через AI: {str(retry_error)}", exc_info=True)
                ai_generation_warning = str(retry_error)
                fallback_context = request.context.strip() if request.context else ""
                ai_description = (
                    "ИИ временно недоступен, использован шаблон оформления нарушения. "
                    f"Заявленный тип нарушения: {request.violation_type}. "
                    f"Оборудование: {equipment.equipment_type} (паспорт: {equipment.passport_number})."
                )
                if fallback_context:
                    ai_description += f" Контекст: {fallback_context}."
        
        # РџСЂРѕРІРµСЂСЏРµРј, С‡С‚Рѕ AI РІРµСЂРЅСѓР» РѕРїРёСЃР°РЅРёРµ
        if not ai_description or not ai_description.strip():
            logger.error("AI РІРµСЂРЅСѓР» РїСѓСЃС‚РѕРµ РѕРїРёСЃР°РЅРёРµ")
            raise HTTPException(
                status_code=500,
                detail="AI РІРµСЂРЅСѓР» РїСѓСЃС‚РѕРµ РѕРїРёСЃР°РЅРёРµ РЅР°СЂСѓС€РµРЅРёСЏ. РџРѕРїСЂРѕР±СѓР№С‚Рµ РµС‰Рµ СЂР°Р· РёР»Рё СѓРІРµР»РёС‡СЊС‚Рµ Р»РёРјРёС‚ С‚РѕРєРµРЅРѕРІ."
            )
        
        # Р›РѕРіРёСЂСѓРµРј РєРѕРЅС‚РµРєСЃС‚, РєРѕС‚РѕСЂС‹Р№ Р±С‹Р» РїРµСЂРµРґР°РЅ РР
        logger.info(f"=== РљРћРќРўР•РљРЎРў Р”Р›РЇ РР ===")
        logger.info(f"РўРёРї РЅР°СЂСѓС€РµРЅРёСЏ: {request.violation_type}")
        logger.info(f"РћР±РѕСЂСѓРґРѕРІР°РЅРёРµ: {equipment.equipment_type} (ID: {equipment.id})")
        logger.info(f"РќР°Р№РґРµРЅРѕ РґРѕРєСѓРјРµРЅС‚РѕРІ РІ Р±Р°Р·Рµ Р·РЅР°РЅРёР№: {len(used_documents)}")
        if used_documents:
            logger.info("РСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹Рµ РґРѕРєСѓРјРµРЅС‚С‹:")
            for doc in used_documents:
                logger.info(f"  - {doc['document_type']}: {doc['title']} (РїСѓРЅРєС‚: {doc.get('clause_number', 'РЅ/Рґ')})")
        else:
            logger.warning("вљ пёЏ Р”РѕРєСѓРјРµРЅС‚С‹ РІ Р±Р°Р·Рµ Р·РЅР°РЅРёР№ РќР• РќРђР™Р”Р•РќР«!")
        logger.info(f"Р”Р»РёРЅР° РєРѕРЅС‚РµРєСЃС‚Р° Р±Р°Р·С‹ Р·РЅР°РЅРёР№: {len(knowledge_context)} СЃРёРјРІРѕР»РѕРІ")
        logger.info("=== РљРћРќР•Р¦ РљРћРќРўР•РљРЎРўРђ ===")
        
        logger.info("РџР°СЂСЃРёРЅРі РѕС‚РІРµС‚Р° AI")
        
        # РџР°СЂСЃРёРј РѕС‚РІРµС‚ AI РґР»СЏ РёР·РІР»РµС‡РµРЅРёСЏ РѕРїРёСЃР°РЅРёСЏ, РїСѓРЅРєС‚РѕРІ Р¤РќРџ/Р“РћРЎРў Рё СЃСЂРѕРєР°
        description = ai_description.strip()
        fnp_clause = None
        gost_clause = None
        deadline_days = 30  # РџРѕ СѓРјРѕР»С‡Р°РЅРёСЋ 30 РґРЅРµР№
        severity = "medium"  # РџРѕ СѓРјРѕР»С‡Р°РЅРёСЋ СЃСЂРµРґРЅСЏСЏ РєСЂРёС‚РёС‡РЅРѕСЃС‚СЊ
        
        # РџС‹С‚Р°РµРјСЃСЏ РёР·РІР»РµС‡СЊ СЃС‚СЂСѓРєС‚СѓСЂРёСЂРѕРІР°РЅРЅС‹Рµ РґР°РЅРЅС‹Рµ РёР· РѕС‚РІРµС‚Р°
        try:
            lines = ai_description.split('\n')
            for i, line in enumerate(lines):
                line = line.strip()
                if line.startswith('РћРџРРЎРђРќРР•:') or line.startswith('РћРџРРЎРђРќРР•'):
                    # Р‘РµСЂРµРј РѕРїРёСЃР°РЅРёРµ РґРѕ СЃР»РµРґСѓСЋС‰РµРіРѕ Р·Р°РіРѕР»РѕРІРєР°
                    desc_lines = []
                    for j in range(i + 1, len(lines)):
                        if lines[j].strip().startswith(('Р¤РќРџ:', 'Р“РћРЎРў:', 'РЎР РћРљ_Р”РќР•Р™:', 'Р¤РќРџ', 'Р“РћРЎРў', 'РЎР РћРљ_Р”РќР•Р™')):
                            break
                        if lines[j].strip():
                            desc_lines.append(lines[j].strip())
                    if desc_lines:
                        description = ' '.join(desc_lines)
                elif line.startswith('Р¤РќРџ:') or line.startswith('Р¤РќРџ'):
                    fnp_text = line.split(':', 1)[-1].strip() if ':' in line else line.replace('Р¤РќРџ', '').strip()
                    if fnp_text and fnp_text.lower() not in ['РЅРµ РїСЂРёРјРµРЅРёРјРѕ', 'РЅРµ РїСЂРёРјРµРЅРёРј', 'РЅ/Рґ', 'РЅ/Р°']:
                        fnp_clause = fnp_text
                elif line.startswith('Р“РћРЎРў:') or line.startswith('Р“РћРЎРў'):
                    gost_text = line.split(':', 1)[-1].strip() if ':' in line else line.replace('Р“РћРЎРў', '').strip()
                    if gost_text and gost_text.lower() not in ['РЅРµ РїСЂРёРјРµРЅРёРјРѕ', 'РЅРµ РїСЂРёРјРµРЅРёРј', 'РЅ/Рґ', 'РЅ/Р°']:
                        gost_clause = gost_text
                elif line.startswith('РЎР РћРљ_Р”РќР•Р™:') or line.startswith('РЎР РћРљ_Р”РќР•Р™') or 'РЎР РћРљ' in line.upper():
                    days_text = line.split(':', 1)[-1].strip() if ':' in line else line
                    # РР·РІР»РµРєР°РµРј С‡РёСЃР»Рѕ РёР· СЃС‚СЂРѕРєРё
                    days_match = re.search(r'\d+', days_text)
                    if days_match:
                        deadline_days = int(days_match.group())
                        # РћРїСЂРµРґРµР»СЏРµРј РєСЂРёС‚РёС‡РЅРѕСЃС‚СЊ РЅР° РѕСЃРЅРѕРІРµ СЃСЂРѕРєР°
                        if deadline_days <= 7:
                            severity = "critical"
                        elif deadline_days <= 15:
                            severity = "high"
                        elif deadline_days <= 30:
                            severity = "medium"
                        else:
                            severity = "low"
        except Exception as parse_error:
            logger.warning(f"РћС€РёР±РєР° РїР°СЂСЃРёРЅРіР° РѕС‚РІРµС‚Р° AI: {parse_error}. РСЃРїРѕР»СЊР·СѓРµРј РІРµСЃСЊ РѕС‚РІРµС‚ РєР°Рє РѕРїРёСЃР°РЅРёРµ.")
            # Р•СЃР»Рё РЅРµ СѓРґР°Р»РѕСЃСЊ СЂР°СЃРїР°СЂСЃРёС‚СЊ, РёСЃРїРѕР»СЊР·СѓРµРј РІРµСЃСЊ РѕС‚РІРµС‚ РєР°Рє РѕРїРёСЃР°РЅРёРµ
        
        # Р’С‹С‡РёСЃР»СЏРµРј РґР°С‚Сѓ РґРµРґР»Р°Р№РЅР°
        deadline = datetime.utcnow() + timedelta(days=deadline_days) if deadline_days > 0 else None
        
        if not has_knowledge_sources:
            fnp_clause = None
            gost_clause = None
            if "В базе знаний не найдено релевантных нормативов" not in description:
                description = (
                    "В базе знаний не найдено релевантных нормативов. "
                    "Уточните запрос. " + description
                )

        if ai_generation_warning and "ИИ временно недоступен" not in description:
            description = "ИИ временно недоступен. Нарушение создано по шаблону. " + description

        logger.info(f"РР·РІР»РµС‡РµРЅРѕ: РѕРїРёСЃР°РЅРёРµ={len(description)} СЃРёРјРІРѕР»РѕРІ, Р¤РќРџ={fnp_clause}, Р“РћРЎРў={gost_clause}, СЃСЂРѕРє={deadline_days} РґРЅРµР№, РєСЂРёС‚РёС‡РЅРѕСЃС‚СЊ={severity}")
        logger.info("РЎРѕР·РґР°РЅРёРµ РЅР°СЂСѓС€РµРЅРёСЏ РІ Р±Р°Р·Рµ РґР°РЅРЅС‹С…")
        
        # РЎРѕР·РґР°РЅРёРµ РЅР°СЂСѓС€РµРЅРёСЏ
        new_violation = Violation(
            inspection_id=request.inspection_id,
            equipment_id=request.equipment_id,
            source="ai_fallback" if ai_generation_warning else "ai",
            description=description,
            fnp_clause=fnp_clause,
            gost_clause=gost_clause,
            severity=severity,
            deadline=deadline,
            status="open",
            created_by=current_user.id
        )
        if deadline:
            new_violation.deadline_source = "ai"
        db.add(new_violation)
        await db.flush()
        
        logger.info(f"РќР°СЂСѓС€РµРЅРёРµ СЃРѕР·РґР°РЅРѕ СЃ ID {new_violation.id}")
        
        # Р›РѕРіРёСЂРѕРІР°РЅРёРµ
        activity = UserActivity(
            user_id=current_user.id,
            action_type="create",
            entity_type="violation",
            entity_id=new_violation.id,
            description=f"AI-generated violation for equipment {request.equipment_id}"
        )
        db.add(activity)

        await log_audit_event(
            db,
            entity_type="violation",
            entity_id=new_violation.id,
            action="CREATE",
            field_changes={
                "status": {"old": None, "new": new_violation.status},
                "severity": {"old": None, "new": new_violation.severity},
                "deadline": {"old": None, "new": new_violation.deadline.isoformat() if new_violation.deadline else None},
                "description": {"old": None, "new": new_violation.description},
                "source": {"old": None, "new": "ai"},
            },
            performed_by=current_user.id,
            source="ai",
            trace_id=getattr(http_request.state, "trace_id", None),
        )

        await db.commit()
        await db.refresh(new_violation)
        new_violation.equipment = equipment
        
        logger.info(f"РќР°СЂСѓС€РµРЅРёРµ СѓСЃРїРµС€РЅРѕ СЃРѕС…СЂР°РЅРµРЅРѕ, РІРѕР·РІСЂР°С‰Р°РµРј РѕС‚РІРµС‚")
        
        violation_response = _violation_to_response(new_violation)
        
        response = AIGenerateViolationResponse(
            violation=violation_response,
            used_documents=used_documents
        )
        
        logger.info(f"Р’РѕР·РІСЂР°С‰Р°РµРј РѕС‚РІРµС‚ СЃ РЅР°СЂСѓС€РµРЅРёРµРј ID {new_violation.id}, РёСЃРїРѕР»СЊР·РѕРІР°РЅРѕ РґРѕРєСѓРјРµРЅС‚РѕРІ: {len(used_documents)}")
        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"РќРµРѕР¶РёРґР°РЅРЅР°СЏ РѕС€РёР±РєР° РїСЂРё РіРµРЅРµСЂР°С†РёРё РЅР°СЂСѓС€РµРЅРёСЏ: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"РћС€РёР±РєР° РіРµРЅРµСЂР°С†РёРё РЅР°СЂСѓС€РµРЅРёСЏ: {str(e)}")

@router.put("/{violation_id}", response_model=ViolationResponse)
async def update_violation(
    violation_id: int,
    violation_data: ViolationUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РћР±РЅРѕРІРёС‚СЊ РЅР°СЂСѓС€РµРЅРёРµ"""
    await require_permission(current_user, "violations:update", db)
    
    result = await db.execute(select(Violation).where(Violation.id == violation_id))
    violation = result.scalar_one_or_none()
    
    if not violation:
        raise HTTPException(status_code=404, detail="Violation not found")
    
    before_state = {
        "status": violation.status,
        "deadline": violation.deadline.isoformat() if violation.deadline else None,
        "severity": violation.severity,
        "description": violation.description,
        "criticality_level": violation.criticality_level,
        "fnp_clause": violation.fnp_clause,
        "gost_clause": violation.gost_clause,
    }

    update_data = violation_data.dict(exclude_unset=True)
    recalc_sla = False
    for field, value in update_data.items():
        setattr(violation, field, value)

    if "deadline" in update_data:
        if update_data["deadline"] is not None:
            violation.deadline_source = "manual"
            violation.deadline_rule_id = None
        else:
            violation.deadline_source = None
            violation.deadline_rule_id = None

    if ("severity" in update_data or "violation_type" in update_data) and violation.deadline_source in ["sla", "sla_default"] and "deadline" not in update_data:
        recalc_sla = True

    if violation_data.status == "resolved" and not violation.resolved_at:
        violation.resolved_at = datetime.utcnow()
        violation.resolved_by = current_user.id
        violation.is_overdue = False
        violation.overdue_at = None
    elif violation_data.status == "open":
        violation.resolved_at = None
        violation.resolved_by = None

    violation.updated_at = datetime.utcnow()

    if recalc_sla:
        violation.deadline = None
        violation.deadline_source = None
        violation.deadline_rule_id = None

    if violation.status != "resolved" and not violation.deadline:
        await _apply_sla_deadline(db, violation)

    after_state = {
        "status": violation.status,
        "deadline": violation.deadline.isoformat() if violation.deadline else None,
        "severity": violation.severity,
        "description": violation.description,
        "criticality_level": violation.criticality_level,
        "fnp_clause": violation.fnp_clause,
        "gost_clause": violation.gost_clause,
    }
    field_changes = build_field_changes(before_state, after_state)
    audit_action = "STATUS_CHANGE" if "status" in field_changes else "UPDATE"
    
    # Р›РѕРіРёСЂРѕРІР°РЅРёРµ
    activity = UserActivity(
        user_id=current_user.id,
        action_type="update",
        entity_type="violation",
        entity_id=violation.id,
        description=f"Updated violation {violation.id}"
    )
    db.add(activity)

    if field_changes:
        await log_audit_event(
            db,
            entity_type="violation",
            entity_id=violation.id,
            action=audit_action,
            field_changes=field_changes,
            performed_by=current_user.id,
            source="ui",
            trace_id=getattr(request.state, "trace_id", None),
        )

    await db.commit()
    result = await db.execute(
        select(Violation)
        .options(selectinload(Violation.equipment))
        .where(Violation.id == violation.id)
    )
    updated_violation = result.scalar_one()
    
    return _violation_to_response(updated_violation)

@router.delete("/{violation_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_violation(
    violation_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """РЈРґР°Р»РёС‚СЊ РЅР°СЂСѓС€РµРЅРёРµ"""
    await require_permission(current_user, "violations:delete", db)
    
    result = await db.execute(select(Violation).where(Violation.id == violation_id))
    violation = result.scalar_one_or_none()
    
    if not violation:
        raise HTTPException(status_code=404, detail="Violation not found")
    
    # Р›РѕРіРёСЂРѕРІР°РЅРёРµ
    activity = UserActivity(
        user_id=current_user.id,
        action_type="delete",
        entity_type="violation",
        entity_id=violation.id,
        description=f"Deleted violation {violation.id}"
    )
    db.add(activity)
    
    await db.delete(violation)
    await db.commit()
    return None

